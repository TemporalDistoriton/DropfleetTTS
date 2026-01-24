import fitz  # PyMuPDF
from PIL import Image
import io
import re
import os
import numpy as np
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def extract_ship_name(page):
    """
    Extract the ship name and type from the top of the page.
    Returns the ship name or a default name if not found.
    """
    # Get text from the page
    text = page.get_text("text")
    
    # Split into lines and get the first few lines
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    if len(lines) >= 3:
        # Line 0 contains ship name (e.g., "Santiago")
        # Line 2 contains ship type (e.g., "Corvette")
        
        ship_name = lines[0].strip()
        ship_type = lines[2].strip()
        
        # Combine ship name and type
        full_name = ship_name + " " + ship_type
        
        # Clean up the full name (remove special characters)
        full_name = re.sub(r'[<>:"/\\|?*]', '', full_name)
        
        return full_name
    
    return "Unknown_Ship"

def extract_ship_data(page):
    """
    Extract all ship data from a page for Excel export.
    
    Returns:
        dict: Dictionary with ship stats, or None if no ship data found
    """
    text = page.get_text("text")
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    # Check if this is a valid ship page (has enough lines)
    if len(lines) < 10:
        return None
    
    try:
        # Extract basic info
        ship_name = lines[0].strip()
        
        # Extract points (look for "XX pts" pattern)
        points = None
        for line in lines[:5]:
            points_match = re.search(r'(\d+)\s*pts', line, re.IGNORECASE)
            if points_match:
                points = int(points_match.group(1))
                break
        
        # Extract ship type (usually line 2)
        ship_type = lines[2].strip() if len(lines) > 2 else ""
        
        # Extract size (look for "L/XXmm" or similar pattern)
        size = None
        for line in lines[:10]:
            size_match = re.search(r'[A-Z]/(\d+)mm', line)
            if size_match:
                size = int(size_match.group(1))
                break
        
        # Find the stats table (look for "Thrust" header)
        thrust_index = None
        for i, line in enumerate(lines):
            if 'Thrust' in line and 'Scan' in line:
                thrust_index = i
                break
        
        if thrust_index is None:
            return None
        
        # The stats should be in the line after the header
        # Format: Thrust Scan Sig Hull ES KS BS G Special
        stats_line_index = thrust_index + 1
        
        # Get the stats line
        if stats_line_index >= len(lines):
            return None
        
        # Parse stats - they might be spread across multiple lines
        # Let's collect all lines that look like stats
        stats_text = ""
        for i in range(stats_line_index, min(stats_line_index + 5, len(lines))):
            stats_text += " " + lines[i]
        
        # Extract individual stats using regex
        # Thrust: digits followed by "
        thrust_match = re.search(r'(\d+)"', stats_text)
        thrust = thrust_match.group(1) if thrust_match else ""
        
        # After first match, look for more
        remaining = stats_text
        if thrust_match:
            remaining = stats_text[thrust_match.end():]
        
        # Scan: next digits followed by "
        scan_match = re.search(r'(\d+)"', remaining)
        scan = scan_match.group(1) if scan_match else ""
        
        if scan_match:
            remaining = remaining[scan_match.end():]
        
        # Sig: next digits followed by "
        sig_match = re.search(r'(\d+)"', remaining)
        sig = sig_match.group(1) if sig_match else ""
        
        if sig_match:
            remaining = remaining[sig_match.end():]
        
        # Hull: just a number
        hull_match = re.search(r'\b(\d+)\b', remaining)
        hull = hull_match.group(1) if hull_match else ""
        
        if hull_match:
            remaining = remaining[hull_match.end():]
        
        # ES: number followed by +
        es_match = re.search(r'(\d+)\+', remaining)
        es = es_match.group(1) if es_match else ""
        
        if es_match:
            remaining = remaining[es_match.end():]
        
        # KS: number followed by +
        ks_match = re.search(r'(\d+)\+', remaining)
        ks = ks_match.group(1) if ks_match else ""
        
        if ks_match:
            remaining = remaining[ks_match.end():]
        
        # BS: could be - or a number
        bs_match = re.search(r'[-\d+]+', remaining)
        bs = bs_match.group(0).strip() if bs_match else ""
        
        if bs_match:
            remaining = remaining[bs_match.end():]
        
        # G: pattern like "2-4" or a single number
        g_match = re.search(r'(\d+-?\d*)', remaining)
        g = g_match.group(1) if g_match else ""
        
        # Special: everything else after the G value
        special = ""
        if g_match:
            special_start = remaining[g_match.end():].strip()
            # Clean up the special text
            special = special_start.split('\n')[0] if special_start else ""
        
        return {
            'name': ship_name,
            'points': points if points else "",
            'type': ship_type,
            'size': size if size else "",
            'thrust': thrust,
            'scan': scan,
            'sig': sig,
            'hull': hull,
            'es': es,
            'ks': ks,
            'bs': bs,
            'g': g,
            'special': special.strip()
        }
        
    except Exception as e:
        print(f"    Error extracting ship data: {str(e)}")
        return None

def get_crop_rect(page, search_text="Famous Ships of the class", crop_whitespace=True):
    """
    Find the text on the page and/or detect content boundaries to crop whitespace.
    
    Args:
        page: PyMuPDF page object
        search_text: Text to search for cropping
        crop_whitespace: Whether to crop empty whitespace at bottom
    
    Returns:
        fitz.Rect: Clipping rectangle, or None if no cropping needed
    """
    page_rect = page.rect
    crop_y = page_rect.height  # Default to full page height
    
    # First, check for the search text
    text_instances = page.search_for(search_text)
    if text_instances:
        # Get the y-coordinate where the text starts
        crop_y = text_instances[0].y0
    
    # If crop_whitespace is enabled, find the actual content boundary
    if crop_whitespace:
        # Get all text blocks with their positions
        blocks = page.get_text("dict")["blocks"]
        
        if blocks:
            # Find the lowest y-coordinate of any content
            max_y = 0
            for block in blocks:
                if "bbox" in block:
                    block_bottom = block["bbox"][3]  # y1 coordinate
                    max_y = max(max_y, block_bottom)
            
            # Add a small margin (e.g., 20 points) below the last content
            content_bottom = max_y + 20
            
            # Use the smaller of the two crop points
            crop_y = min(crop_y, content_bottom)
    
    # Only return a crop rectangle if we're actually cropping
    if crop_y < page_rect.height:
        clip_rect = fitz.Rect(0, 0, page_rect.width, crop_y)
        return clip_rect
    
    return None

def remove_background(image, threshold=240, black_threshold=15):
    """
    Remove white and black backgrounds from an image and make them transparent.
    
    Args:
        image: PIL Image object
        threshold: Pixel values above this are considered white (default: 240)
        black_threshold: Pixel values below this are considered black (default: 15)
    
    Returns:
        PIL Image with transparent background
    """
    # Convert to RGBA if not already
    if image.mode != 'RGBA':
        image = image.convert('RGBA')
    
    # Convert to numpy array for easier manipulation
    data = np.array(image)
    
    # Get RGB channels
    r, g, b, a = data[:,:,0], data[:,:,1], data[:,:,2], data[:,:,3]
    
    # Create mask for white pixels (all RGB values above threshold)
    white_mask = (r > threshold) & (g > threshold) & (b > threshold)
    
    # Create mask for black pixels (all RGB values below black_threshold)
    black_mask = (r < black_threshold) & (g < black_threshold) & (b < black_threshold)
    
    # Combine masks
    background_mask = white_mask | black_mask
    
    # Set alpha channel to 0 (transparent) for background pixels
    data[background_mask, 3] = 0
    
    # Convert back to PIL Image
    return Image.fromarray(data)

def extract_ship_image(page, ship_name, output_folder, dpi=300):
    """
    Extract the ship image from the page, ignoring the background image (2480x3507).
    Remove its background and save it with transparency.
    
    Args:
        page: PyMuPDF page object
        ship_name: Name for the output file
        output_folder: Folder to save the image
        dpi: Resolution for extraction
    
    Returns:
        bool: True if image was extracted successfully
    """
    # Get all images on the page
    image_list = page.get_images(full=True)
    
    if not image_list:
        return False
    
    # Find the largest non-background image
    largest_image = None
    largest_area = 0
    largest_xref = None
    
    for img in image_list:
        xref = img[0]
        
        # Get the actual image dimensions
        try:
            base_image = page.parent.extract_image(xref)
            width = base_image["width"]
            height = base_image["height"]
            
            # Skip the background image (2480x3507)
            if width == 2480 and height == 3507:
                continue
            
            # Calculate area
            area = width * height
            
            # Track the largest non-background image
            if area > largest_area:
                largest_area = area
                largest_xref = xref
                
        except Exception as e:
            continue
    
    if not largest_xref:
        return False
    
    # Extract the ship image
    try:
        base_image = page.parent.extract_image(largest_xref)
        image_bytes = base_image["image"]
        
        # Load image with PIL
        image = Image.open(io.BytesIO(image_bytes))
        
        # Remove background and make transparent
        image_transparent = remove_background(image)
        
        # Save the image
        output_filename = f"{ship_name}_ModelImage.png"
        output_path = os.path.join(output_folder, output_filename)
        image_transparent.save(output_path, "PNG")
        
        return True
        
    except Exception as e:
        print(f"    Warning: Could not extract ship image - {str(e)}")
        return False

def pdf_to_png(pdf_path, output_folder="output", dpi=300, crop_text="Famous Ships of the class", 
               crop_whitespace=True, extract_model=True, ship_data_list=None):
    """
    Convert each page of a PDF to PNG with ship name in filename.
    Crops content at specified text and/or removes bottom whitespace.
    Optionally extracts ship model images with transparent backgrounds.
    
    Args:
        pdf_path: Path to the input PDF file
        output_folder: Folder to save PNG files (default: 'output')
        dpi: Resolution for the output images (default: 300)
        crop_text: Text to search for cropping (default: 'Famous Ships of the class')
        crop_whitespace: Whether to crop empty whitespace at bottom (default: True)
        extract_model: Whether to extract ship model images (default: True)
        ship_data_list: List to append ship data to for Excel export
    """
    # Create output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Open the PDF
    pdf_document = fitz.open(pdf_path)
    
    print(f"Processing PDF: {pdf_path}")
    print(f"Total pages: {len(pdf_document)}")
    
    # Process each page
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]
        
        # Extract ship data for Excel
        ship_data = extract_ship_data(page)
        if ship_data and ship_data_list is not None:
            ship_data_list.append(ship_data)
        elif ship_data_list is not None:
            # No ship data found
            ship_data_list.append({'name': 'no ship'})
        
        # Extract ship name from the page
        ship_name = extract_ship_name(page)
        
        # Create the output filename for card front
        output_filename = f"{ship_name}_CardFrontImage.png"
        output_path = os.path.join(output_folder, output_filename)
        
        # Calculate zoom factor based on DPI (72 is the default PDF DPI)
        zoom = dpi / 72
        mat = fitz.Matrix(zoom, zoom)
        
        # Check if we need to crop the page
        clip_rect = get_crop_rect(page, crop_text, crop_whitespace)
        
        if clip_rect:
            # Render only the cropped portion
            pix = page.get_pixmap(matrix=mat, clip=clip_rect)
            print(f"Page {page_num + 1}: Cropped - Saved as '{output_filename}'")
        else:
            # Render the full page
            pix = page.get_pixmap(matrix=mat)
            print(f"Page {page_num + 1}: Full page - Saved as '{output_filename}'")
        
        # Save card front as PNG
        pix.save(output_path)
        
        # Extract ship model image if requested
        if extract_model:
            success = extract_ship_image(page, ship_name, output_folder, dpi)
            if success:
                print(f"    ✓ Extracted model image")
            else:
                print(f"    ✗ No model image found")
    
    pdf_document.close()
    print(f"\nConversion complete! Files saved in '{output_folder}' folder.")

def create_excel_file(ship_data_list, output_file="ship_data.xlsx"):
    """
    Create an Excel file with all ship data.
    
    Args:
        ship_data_list: List of ship data dictionaries
        output_file: Name of the output Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ship Data"
    
    # Define headers
    headers = ['Name', 'Points', 'Type', 'Size', 'Thrust', 'Scan', 'Sig', 
               'Hull', 'ES', 'KS', 'BS', 'G', 'Special']
    
    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, ship in enumerate(ship_data_list, start=2):
        if ship.get('name') == 'no ship':
            ws.cell(row=row_idx, column=1, value='no ship')
        else:
            ws.cell(row=row_idx, column=1, value=ship.get('name', ''))
            ws.cell(row=row_idx, column=2, value=ship.get('points', ''))
            ws.cell(row=row_idx, column=3, value=ship.get('type', ''))
            ws.cell(row=row_idx, column=4, value=ship.get('size', ''))
            ws.cell(row=row_idx, column=5, value=ship.get('thrust', ''))
            ws.cell(row=row_idx, column=6, value=ship.get('scan', ''))
            ws.cell(row=row_idx, column=7, value=ship.get('sig', ''))
            ws.cell(row=row_idx, column=8, value=ship.get('hull', ''))
            ws.cell(row=row_idx, column=9, value=ship.get('es', ''))
            ws.cell(row=row_idx, column=10, value=ship.get('ks', ''))
            ws.cell(row=row_idx, column=11, value=ship.get('bs', ''))
            ws.cell(row=row_idx, column=12, value=ship.get('g', ''))
            ws.cell(row=row_idx, column=13, value=ship.get('special', ''))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved: {output_file}")

# Example usage
if __name__ == "__main__":
    import sys
    
    # DPI setting (can be adjusted here if needed)
    dpi = 300  # Higher DPI = better quality but larger files
    
    # Find all PDF files in the current directory
    pdf_files = glob.glob("*.pdf")
    
    if not pdf_files:
        print("No PDF files found in the current directory!")
        print("Please make sure there are PDF files in the same folder as this script.")
        sys.exit(1)
    
    print(f"Found {len(pdf_files)} PDF file(s) to process:")
    for pdf in pdf_files:
        print(f"  - {pdf}")
    print()
    
    # List to collect all ship data for Excel
    all_ship_data = []
    
    # Process each PDF file
    for pdf_file in pdf_files:
        # Create output folder name from PDF filename (without extension)
        pdf_basename = os.path.splitext(os.path.basename(pdf_file))[0]
        output_folder = pdf_basename
        
        try:
            pdf_to_png(pdf_file, output_folder, dpi, ship_data_list=all_ship_data)
            print()
        except Exception as e:
            print(f"Error processing {pdf_file}: {str(e)}")
            print()
    
    # Create Excel file with all ship data
    if all_ship_data:
        create_excel_file(all_ship_data, "ship_data.xlsx")
    
    print("=" * 60)
    print("All PDFs processed!")