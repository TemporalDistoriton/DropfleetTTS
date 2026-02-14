"""
Replace Steam Assets with GitHub Hosted Assets in a TTS Save
--------------------------------------------------------------
Reads a TTS save JSON and the Steam_Assets_Table.xlsx spreadsheet.
For each GUID, finds the matching object in the save, checks if it
still has Steam-hosted assets, looks for a GitHub-hosted replacement,
and swaps the URL if found.

The GitHub assets are expected at:
  https://raw.githubusercontent.com/TemporalDistoriton/DropfleetTTS/main/EXTRACTOR/downloaded_assets/
  Organised in folders named: {GUID} - {Object Name}/

Usage:
    python replace_steam_with_github.py <tts_save.json> <spreadsheet.xlsx> [output.json]

Requirements:
    pip install openpyxl requests
"""

import os
import sys
import re
import requests
import time
from urllib.parse import urlparse, quote
from openpyxl import load_workbook

# --- Configuration ---
GITHUB_REPO = "TemporalDistoriton/DropfleetTTS"
GITHUB_BRANCH = "main"
GITHUB_BASE_PATH = "EXTRACTOR/downloaded_assets"
RAW_BASE_URL = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{GITHUB_BASE_PATH}"
TREE_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/git/trees/{GITHUB_BRANCH}?recursive=1"

STEAM_DOMAIN = "steamusercontent-a.akamaihd.net"


# --- Helpers (must match the download script exactly) ---

def sanitize_folder_name(name):
    """Same logic as the download script."""
    sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name)
    sanitized = sanitized.strip(' .')
    sanitized = re.sub(r'_+', '_', sanitized)
    return sanitized if sanitized else "unnamed_object"


def get_url_hash(url):
    """Same logic as the download script — last 16 chars of cleaned path."""
    return urlparse(url).path.strip("/").replace("/", "_")[-16:]


# --- Core functions ---

def read_excel(filepath):
    """Read spreadsheet and return list of (object_name, guid, [steam_urls])."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    objects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        guid = str(row[1]).strip() if row[1] else "no_guid"
        links = [
            str(cell).strip()
            for cell in row[2:]
            if cell and STEAM_DOMAIN in str(cell)
        ]
        if links:
            objects.append((name, guid, links))
    wb.close()
    return objects


def fetch_github_tree(token=None):
    """
    Fetch the entire repo file tree in a single API call.
    Returns a dict mapping: "folder_name/filename" -> True
    for every file under the downloaded_assets path.
    """
    headers = {"Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"token {token}"

    print("Fetching GitHub repository file tree...")
    resp = requests.get(TREE_API_URL, headers=headers, timeout=30)
    resp.raise_for_status()
    tree_data = resp.json()

    if tree_data.get("truncated"):
        print("  ⚠ Warning: Repository tree was truncated (very large repo).")
        print("    Some assets may not be found. Consider using a GitHub token.")

    prefix = GITHUB_BASE_PATH + "/"
    files = {}
    for item in tree_data.get("tree", []):
        if item["type"] == "blob" and item["path"].startswith(prefix):
            relative = item["path"][len(prefix):]  # "GUID - Name/asset_01_hash.png"
            files[relative] = True

    print(f"  Found {len(files)} asset files on GitHub.\n")
    return files


def build_replacement_map(objects, github_files):
    """
    Match each Steam URL to its GitHub counterpart.

    For each object, reconstruct the expected folder name, then match
    Steam URLs to GitHub filenames via the url_hash embedded in each filename.
    """
    replacements = {}
    matched_objects = 0
    unmatched_objects = []
    unmatched_urls = []

    for name, guid, steam_urls in objects:
        folder_name = sanitize_folder_name(f"{guid} - {name}")

        # Gather all GitHub files in this folder and index by url_hash
        hash_to_github_file = {}
        for rel_path in github_files:
            if not rel_path.startswith(folder_name + "/"):
                continue
            filename = rel_path.split("/", 1)[1]
            # Extract hash from "asset_01_XXXXXXXXXXXXXXXX.ext"
            match = re.match(r'asset_\d+_(.+)\.\w+$', filename)
            if match:
                hash_to_github_file[match.group(1)] = rel_path

        if not hash_to_github_file:
            unmatched_objects.append((guid, name, folder_name))
            continue

        obj_matched = 0
        for url in steam_urls:
            url_hash = get_url_hash(url)
            if url_hash in hash_to_github_file:
                rel_path = hash_to_github_file[url_hash]
                # URL-encode each path segment separately to handle spaces/special chars
                encoded_path = "/".join(quote(seg, safe="") for seg in rel_path.split("/"))
                github_url = f"{RAW_BASE_URL}/{encoded_path}"
                replacements[url] = github_url
                obj_matched += 1
            else:
                unmatched_urls.append((guid, name, url))

        if obj_matched > 0:
            matched_objects += 1

    return replacements, matched_objects, unmatched_objects, unmatched_urls


def apply_replacements(save_text, replacements):
    """Replace all Steam URLs with their GitHub counterparts in the save text."""
    replaced_count = 0
    # Sort by longest URL first to avoid partial replacements
    for steam_url in sorted(replacements, key=len, reverse=True):
        github_url = replacements[steam_url]
        count = save_text.count(steam_url)
        if count > 0:
            save_text = save_text.replace(steam_url, github_url)
            replaced_count += count
    return save_text, replaced_count


def main():
    if len(sys.argv) < 3:
        print("Usage: python replace_steam_with_github.py <tts_save.json> <spreadsheet.xlsx> [output.json]")
        print("\nOptional env var: GITHUB_TOKEN=ghp_xxx for higher API rate limits")
        sys.exit(1)

    save_path = sys.argv[1]
    excel_path = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else save_path.replace(".json", "_github.json")

    # Avoid overwriting input by accident
    if os.path.abspath(output_path) == os.path.abspath(save_path):
        output_path = save_path.replace(".json", "_github.json")
        print(f"  Output path same as input — writing to: {output_path}")

    for path in [save_path, excel_path]:
        if not os.path.exists(path):
            print(f"Error: File '{path}' not found.")
            sys.exit(1)

    # Optional GitHub token from environment
    github_token = os.environ.get("GITHUB_TOKEN")

    # Step 1: Read spreadsheet
    print(f"Reading spreadsheet: {excel_path}")
    objects = read_excel(excel_path)
    total_urls = sum(len(links) for _, _, links in objects)
    print(f"  {len(objects)} objects with {total_urls} Steam asset URLs.\n")

    # Step 2: Fetch GitHub file tree (single API call)
    github_files = fetch_github_tree(token=github_token)

    # Step 3: Build replacement map
    print("Matching Steam URLs to GitHub assets...")
    replacements, matched_objs, unmatched_objs, unmatched_urls = build_replacement_map(objects, github_files)
    print(f"  {len(replacements)}/{total_urls} URLs matched to GitHub assets.")
    print(f"  {matched_objs}/{len(objects)} objects have at least one match.\n")

    if unmatched_objs:
        print(f"  ⚠ {len(unmatched_objs)} objects had no GitHub folder:")
        for guid, name, folder in unmatched_objs[:10]:
            print(f"      {guid} - {name}  (expected: {folder}/)")
        if len(unmatched_objs) > 10:
            print(f"      ... and {len(unmatched_objs) - 10} more")
        print()

    if unmatched_urls:
        print(f"  ⚠ {len(unmatched_urls)} individual URLs had no matching GitHub file.")
        for guid, name, url in unmatched_urls[:5]:
            print(f"      {guid} | {url[:80]}...")
        if len(unmatched_urls) > 5:
            print(f"      ... and {len(unmatched_urls) - 5} more")
        print()

    if not replacements:
        print("Nothing to replace — exiting.")
        sys.exit(0)

    # Step 4: Read save file
    print(f"Reading TTS save: {save_path}")
    with open(save_path, "r", encoding="utf-8-sig") as f:
        save_text = f.read()

    steam_count_before = save_text.count(STEAM_DOMAIN)
    print(f"  Steam URLs in save (before): {steam_count_before}\n")

    # Step 5: Apply replacements
    print("Applying replacements...")
    save_text, replaced_count = apply_replacements(save_text, replacements)

    steam_count_after = save_text.count(STEAM_DOMAIN)

    # Step 6: Write output
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(save_text)

    # Summary
    print()
    print("=" * 60)
    print("REPLACEMENT COMPLETE")
    print(f"  Unique URLs replaced:    {len(replacements)}")
    print(f"  Total replacements made: {replaced_count}")
    print(f"  Steam URLs before:       {steam_count_before}")
    print(f"  Steam URLs after:        {steam_count_after}")
    print(f"  Output saved to:         {output_path}")
    print("=" * 60)

    if steam_count_after > 0:
        print(f"\n  Note: {steam_count_after} Steam URLs remain in the save.")
        print("  These may be assets not in the spreadsheet or not yet on GitHub.")

    # Write a log of all replacements for reference
    log_path = output_path.replace(".json", "_replacement_log.txt")
    with open(log_path, "w") as f:
        f.write("STEAM URL → GITHUB URL\n")
        f.write("=" * 120 + "\n\n")
        for steam_url, github_url in sorted(replacements.items()):
            f.write(f"{steam_url}\n  → {github_url}\n\n")
        if unmatched_urls:
            f.write("\n\nUNMATCHED STEAM URLS (no GitHub asset found)\n")
            f.write("=" * 120 + "\n\n")
            for guid, name, url in unmatched_urls:
                f.write(f"  {guid} - {name}: {url}\n")
    print(f"\n  Replacement log: {log_path}")


if __name__ == "__main__":
    main()
