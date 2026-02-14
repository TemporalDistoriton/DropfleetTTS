"""
Download Steam Assets from TTS Save Export
-------------------------------------------
Reads the Steam_Assets_Table.xlsx file and downloads each asset
into a folder named after its parent object.

Usage:
    python download_steam_assets.py Steam_Assets_Table.xlsx [output_directory]

Requirements:
    pip install openpyxl requests
"""

import os
import sys
import re
import time
import requests
from openpyxl import load_workbook
from urllib.parse import urlparse


# --- Configuration ---
REQUEST_TIMEOUT = 30
RETRY_ATTEMPTS = 3
RETRY_DELAY = 2  # seconds between retries
CHUNK_SIZE = 8192

# Content-type to extension mapping for Steam assets
CONTENT_TYPE_MAP = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/gif": ".gif",
    "image/bmp": ".bmp",
    "image/webp": ".webp",
    "application/pdf": ".pdf",
    "audio/ogg": ".ogg",
    "audio/mpeg": ".mp3",
    "audio/wav": ".wav",
    "video/mp4": ".mp4",
    "application/octet-stream": ".asset",  # generic binary fallback
    "text/plain": ".txt",
    "text/html": ".html",
    "application/json": ".json",
    "model/obj": ".obj",
}


def sanitize_folder_name(name):
    """Remove or replace characters that are invalid in folder names."""
    # Replace problematic characters with underscores
    sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name)
    # Strip leading/trailing whitespace and dots
    sanitized = sanitized.strip(' .')
    # Collapse multiple underscores
    sanitized = re.sub(r'_+', '_', sanitized)
    # Fallback if empty
    return sanitized if sanitized else "unnamed_object"


def guess_extension(response):
    """Guess the file extension from the HTTP response headers."""
    content_type = response.headers.get("Content-Type", "").split(";")[0].strip().lower()
    if content_type in CONTENT_TYPE_MAP:
        return CONTENT_TYPE_MAP[content_type]

    # Try Content-Disposition header
    disposition = response.headers.get("Content-Disposition", "")
    if "filename=" in disposition:
        fname = disposition.split("filename=")[-1].strip('" ')
        _, ext = os.path.splitext(fname)
        if ext:
            return ext

    return ".asset"  # generic fallback


def download_file(url, dest_path):
    """Download a file with retries. Returns True on success."""
    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            response = requests.get(url, timeout=REQUEST_TIMEOUT, stream=True)
            response.raise_for_status()

            # Determine extension from response
            ext = guess_extension(response)
            final_path = dest_path + ext

            with open(final_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=CHUNK_SIZE):
                    f.write(chunk)

            size_kb = os.path.getsize(final_path) / 1024
            return True, final_path, size_kb

        except requests.RequestException as e:
            if attempt < RETRY_ATTEMPTS:
                print(f"      Retry {attempt}/{RETRY_ATTEMPTS} after error: {e}")
                time.sleep(RETRY_DELAY)
            else:
                return False, str(e), 0


def read_excel(filepath):
    """Read the Excel file and return a list of (object_name, guid, [links])."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    objects = []
    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        guid = str(row[1]).strip() if row[1] else "no_guid"
        links = [
            str(cell).strip()
            for cell in row[2:]
            if cell and "steamusercontent-a.akamaihd.net" in str(cell)
        ]
        if links:
            objects.append((name, guid, links))

    return objects


def main():
    if len(sys.argv) < 2:
        print("Usage: python download_steam_assets.py <excel_file> [output_dir]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "downloaded_assets"

    if not os.path.exists(excel_path):
        print(f"Error: File '{excel_path}' not found.")
        sys.exit(1)

    # Read data
    print(f"Reading {excel_path}...")
    objects = read_excel(excel_path)
    total_links = sum(len(links) for _, _, links in objects)
    print(f"Found {len(objects)} objects with {total_links} total assets to download.\n")

    # Prepare folders named as "GUID - Object Name"
    folder_names = [
        sanitize_folder_name(f"{guid} - {name}") for name, guid, _ in objects
    ]

    os.makedirs(output_dir, exist_ok=True)

    # Download
    success_count = 0
    fail_count = 0
    failed_items = []

    for idx, ((name, guid, links), folder_name) in enumerate(zip(objects, folder_names)):
        folder_path = os.path.join(output_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        print(f"[{idx + 1}/{len(objects)}] {name} (GUID: {guid}) — {len(links)} asset(s)")

        for link_idx, url in enumerate(links, 1):
            # Use a short hash from URL as base filename to avoid collisions
            url_hash = urlparse(url).path.strip("/").replace("/", "_")[-16:]
            base_name = f"asset_{link_idx:02d}_{url_hash}"
            dest_base = os.path.join(folder_path, base_name)

            ok, result, size_kb = download_file(url, dest_base)
            if ok:
                success_count += 1
                print(f"   ✓ [{link_idx}/{len(links)}] {os.path.basename(result)} ({size_kb:.1f} KB)")
            else:
                fail_count += 1
                failed_items.append((name, guid, url, result))
                print(f"   ✗ [{link_idx}/{len(links)}] FAILED: {result}")

    # Summary
    print("\n" + "=" * 60)
    print(f"DOWNLOAD COMPLETE")
    print(f"  Succeeded: {success_count}")
    print(f"  Failed:    {fail_count}")
    print(f"  Output:    {os.path.abspath(output_dir)}")
    print("=" * 60)

    if failed_items:
        fail_log = os.path.join(output_dir, "_failed_downloads.txt")
        with open(fail_log, "w") as f:
            for name, guid, url, err in failed_items:
                f.write(f"{name} | {guid} | {url} | {err}\n")
        print(f"\nFailed downloads logged to: {fail_log}")


if __name__ == "__main__":
    main()
