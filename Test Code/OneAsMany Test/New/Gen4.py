#!/usr/bin/env python3
"""
TTS Ship Generator v3
=====================
Reads a Tabletop Simulator .json save file (with faction container tiles) and an
Excel spreadsheet of ship data. For each row in the spreadsheet:

1. Clones the template ship model and updates its Lua variables
2. Wraps it in a spawner tile (like the SAVESHIPMODELS tool would)
3. Embeds the spawner tile's JSON into the matching faction container tile
   (matching PDF ID -> "Faction XXX" tile nickname)

The faction container tiles use an embedded Lua data section at the end of their
script to store ship cards, which this tool populates directly.

Usage:
    python tts_ship_generator.py <tts_save.json> <ship_data.xlsx> [output.json]

    The TTS save must contain:
      - One Custom_Model object with SHIP_ID in its LuaScript (the template)
      - Faction container tiles named "Faction BIO", "Faction UCM", etc.
"""

import json
import copy
import re
import sys
import random
import openpyxl


# ─── Helpers ────────────────────────────────────────────────────────────────

def generate_guid(existing_guids: set) -> str:
    """Generate a unique 6-char hex GUID not already in use."""
    while True:
        guid = ''.join(random.choices('0123456789abcdef', k=6))
        if guid not in existing_guids:
            existing_guids.add(guid)
            return guid


def collect_all_guids(save_data: dict) -> set:
    """Walk the entire save and collect every GUID already in use."""
    guids = set()

    def _walk(obj):
        if isinstance(obj, dict):
            if 'GUID' in obj:
                guids.add(obj['GUID'])
            for v in obj.values():
                _walk(v)
        elif isinstance(obj, list):
            for item in obj:
                _walk(item)

    _walk(save_data)
    return guids


def replace_lua_variable(lua: str, var_name: str, new_value, is_string: bool = False) -> str:
    """
    Replace a Lua local variable assignment in the script.
    Uses lambda replacement to avoid regex backreference issues.
    """
    if is_string:
        pattern = rf'(local\s+{re.escape(var_name)}\s*=\s*)"[^"]*"'
        return re.sub(pattern, lambda m: f'{m.group(1)}"{new_value}"', lua, count=1)
    else:
        pattern = rf'(local\s+{re.escape(var_name)}\s*=\s*)\S+'
        return re.sub(pattern, lambda m: f'{m.group(1)}{new_value}', lua, count=1)


def to_num(val):
    """Convert a value to int or float; pass through if already numeric."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val if not isinstance(val, float) or val != int(val) else int(val)
    try:
        s = str(val).strip()
        if '.' in s:
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        return 0


def parse_special_flags(row: dict) -> tuple:
    """
    Parse the Special column text to detect isVectored and isMonitored flags.
    Returns (is_vectored: bool, is_monitored: bool).

    - isVectored  : Special text contains the word 'Vectored' (case-insensitive)
    - isMonitored : Special text contains the word 'Monitor'  (case-insensitive,
                    matches 'Monitor' and 'Monitor,' but NOT 'Monitored' as a
                    different keyword — either way the substring check covers both)
    """
    special = str(row.get('Special', '') or '').lower()
    is_vectored  = 'vectored' in special
    is_monitored = 'monitor'  in special
    return is_vectored, is_monitored


def update_lua_script_state(state_json: str, row: dict) -> str:
    """
    Parse and update the JSON stored in LuaScriptState to keep it in sync
    with the Lua variable changes.
    """
    if not state_json or not state_json.strip():
        return state_json
    try:
        state = json.loads(state_json)
    except json.JSONDecodeError:
        return state_json

    is_vectored, is_monitored = parse_special_flags(row)

    state['shipID']          = row['Name']
    state['Shiphealth']      = to_num(row['Hull'])
    state['ShipbaseSize']    = to_num(row['Size'])
    state['Shipcost']        = to_num(row['Points'])
    state['ShipScan']        = to_num(row['Scan'])
    state['ShipThrust']      = to_num(row['Thrust'])
    state['Signature']       = to_num(row['Sig'])
    state['SHIPCardURL']     = row['Card Image']
    state['FIXED_IMAGE_URL'] = row['Model Image']
    state['baseSignature']   = to_num(row['Sig'])
    state['Vectored']        = is_vectored
    state['Monitor']         = is_monitored

    model_3d_exists = str(row.get('3D MODEL FILE', 'None')).strip() not in ('None', '', 'none')
    state['Model3DAvailable'] = model_3d_exists

    for block_key in ('originalData', 'state'):
        block = state.get(block_key)
        if isinstance(block, dict):
            if 'health' in block:
                hp = to_num(row['Hull'])
                block['health']['current'] = hp
                block['health']['max']     = hp
            if 'base' in block:
                block['base']['size'] = to_num(row['Size'])
            if 'sig' in block:
                block['sig'] = to_num(row['Sig'])

    return json.dumps(state)


# ─── Tonnage classification ────────────────────────────────────────────────

TONNAGE_PATTERNS = [
    ("dreadnought",     "Dreadnaught"),
    ("dreadnaught",     "Dreadnaught"),
    ("super battleship","Super Battleship"),
    ("superbattleship", "Super Battleship"),
    ("supercarrier",    "Supercarrier"),
    ("battleship",      "Battleship"),
    ("battlecruiser",   "Battlecruiser"),
    ("troopship",       "Troopship"),
    ("heavy cruiser",   "Heavy Cruiser"),
    ("light cruiser",   "Light Cruiser"),
    ("cruiser",         "Cruiser"),
    ("runner",          "Runner"),
    ("destroyer",       "Destroyer"),
    ("cutter",          "Cutter"),
    ("monitor",         "Monitor"),
    ("carrier",         "Carrier"),
    ("frigate",         "Frigate"),
    ("lighter",         "Lighter"),
    ("corvette",        "Corvette"),
    ("cell",            "Cell"),
]

SHIPBASE_TO_TONNAGE = {
    20: "Light",
    30: "Light",
    40: "Medium",
    50: "Heavy",
    60: "Colossal",
}


def detect_class_tonnage(name: str, description: str) -> str:
    text = (description.lower() + " " + name.lower())
    for pattern, tonnage in TONNAGE_PATTERNS:
        if pattern in text:
            return tonnage
    return "Other"


def detect_ship_tonnage(base_size: int) -> str:
    return SHIPBASE_TO_TONNAGE.get(base_size, "Light")


# ─── Spawner tile creation ─────────────────────────────────────────────────

SPAWNER_SCRIPT_TEMPLATE = """\
-- SPAWNER TILE: {ship_name}
-- Click the button to spawn the model on this tile

function onLoad()
  local pos = self.getPosition()
  rebuildUI()
end

function rebuildUI()
    local ui = {{
        {{tag='Defaults', children={{
            {{tag='Text', attributes={{color='#cccccc', fontSize='18', alignment='MiddleLeft'}}}},
            {{tag='InputField', attributes={{fontSize='24', preferredHeight='40'}}}},
            {{tag='ToggleButton', attributes={{fontSize='18', preferredHeight='40', colors='#ffcc33|#ffffff|#808080|#606060', selectedBackgroundColor='#dddddd', deselectedBackgroundColor='#999999'}}}},
            {{tag='Button', attributes={{fontSize='12',textColor='#111111', preferredHeight='40', colors='#dddddd|#ffffff|#808080|#f6f6f6'}}}},
            {{tag='Toggle', attributes={{textColor='#cccccc'}}}},
        }}}},

        {{tag='button', attributes={{onClick='ui_createModel',text='Spawn Ship',  colors='#ccccccff|#ffffffff|#404040ff|#808080ff', width='120', height='20', position='0 110 -5', rotation='0 0 180' }}}}
    }}

    self.UI.setXmlTable(ui)
end

function ui_createModel(player, value, id)
    local color = player.color
    local pos = self.getPosition()
    local rot = {{x = 0, y = 0, z = 0}}
    if color == "Blue" then
        rot.y = 180
    end
    spawnObjectJSON({{
        json = objectJSON,
        position = {{x = pos.x, y = pos.y + 2, z = pos.z}},
        rotation = rot,
        callback_function = function(spawned_object)
            Wait.frames(function()
                spawned_object.call("SetPlayerColor", {{playerColor = color}})
            end, 30)
        end
    }})
    broadcastToAll("Spawned: {ship_name} (" .. color .. ")", {{0, 1, 0}})
end

objectJSON = [={delim}[{ship_json}]={delim}]
"""


def pick_lua_delimiter(text: str) -> str:
    level = ''
    while f']={level}]' in text:
        level += '='
    return level


def make_spawner_tile(ship_obj: dict, card_url: str, guid: str) -> dict:
    ship_name = ship_obj.get('Nickname', 'Unknown Ship')
    ship_desc = ship_obj.get('Description', '')
    ship_json = json.dumps(ship_obj, ensure_ascii=False)

    delim = pick_lua_delimiter(ship_json)
    spawner_lua = SPAWNER_SCRIPT_TEMPLATE.format(
        ship_name=ship_name,
        ship_json=ship_json,
        delim=delim,
    )

    return {
        "GUID": guid,
        "Name": "Custom_Tile",
        "Transform": {
            "posX": 0, "posY": 1.0, "posZ": 0,
            "rotX": 0, "rotY": 0, "rotZ": 0,
            "scaleX": 2.0, "scaleY": 1.0, "scaleZ": 2.0,
        },
        "Nickname": ship_name,
        "Description": ship_desc,
        "GMNotes": "",
        "AltLookAngle": {"x": 0.0, "y": 0.0, "z": 0.0},
        "ColorDiffuse": {"r": 1.0, "g": 1.0, "b": 1.0},
        "LayoutGroupSortIndex": 0,
        "Value": 0,
        "Locked": False,
        "Grid": True,
        "Snap": True,
        "IgnoreFoW": False,
        "MeasureMovement": False,
        "DragSelectable": True,
        "Autoraise": True,
        "Sticky": True,
        "Tooltip": True,
        "GridProjection": False,
        "HideWhenFaceDown": False,
        "Hands": False,
        "CustomImage": {
            "ImageURL": card_url,
            "ImageSecondaryURL": card_url,
            "ImageScalar": 1.0,
            "WidthScale": 0.0,
            "CustomTile": {
                "Type": 0,
                "Thickness": 0.1,
                "Stackable": False,
                "Stretch": True,
            },
        },
        "LuaScript": spawner_lua,
        "LuaScriptState": "",
        "XmlUI": "",
    }


# ─── Embedded data generation ──────────────────────────────────────────────

def build_embedded_data(card_entries: list) -> str:
    lines = []
    lines.append("\n--EMBEDDED_DATA_START")
    lines.append("savedShipCards = {")

    for card in card_entries:
        lines.append("  {")
        lines.append(f"    name = [====[{card['name']}]====],")
        lines.append(f"    imageURL = [====[{card['imageURL']}]====],")
        lines.append(f"    tonnage = [====[{card['tonnage']}]====],")
        lines.append(f"    shipTonnage = [====[{card['shipTonnage']}]====],")
        lines.append(f"    baseScale = {card['baseScale']},")
        lines.append(f"    json = [====[{card['json']}]====],")
        lines.append("  },")

    lines.append("}")
    lines.append("")
    lines.append("fresh = false")
    return "\n".join(lines)


# ─── Ship template processing ──────────────────────────────────────────────

def find_template_object(save_data: dict) -> tuple:
    for i, obj in enumerate(save_data.get('ObjectStates', [])):
        if 'SHIP_ID' in obj.get('LuaScript', ''):
            return i, obj
    raise RuntimeError("No template object with SHIP_ID found in the save file.")


def read_ship_rows(xlsx_path: str) -> list:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if row_dict.get('Name') is None:
            continue
        rows.append(row_dict)
    return rows


def create_ship_from_template(template: dict, row: dict, existing_guids: set) -> dict:
    """
    Deep-copy the template and replace all Lua variables + model URLs
    according to one row of Excel data.
    """
    ship = copy.deepcopy(template)

    ship['GUID'] = generate_guid(existing_guids)
    for child in ship.get('ChildObjects', []):
        child['GUID'] = generate_guid(existing_guids)

    ship['Nickname'] = str(row['Name'])
    ship['Description'] = str(row.get('Type', ''))

    model_file = str(row.get('3D MODEL FILE', 'None')).strip()
    texture_file = str(row.get('3D MODEL TEXTURE', 'None')).strip()
    model_3d_exists = model_file not in ('None', '', 'none')

    # Parse isVectored / isMonitored from the Special column text
    is_vectored, is_monitored = parse_special_flags(row)

    lua = ship['LuaScript']
    lua = replace_lua_variable(lua, 'SHIP_ID',         row['Name'],        is_string=True)
    lua = replace_lua_variable(lua, 'Shiphealth',      to_num(row['Hull']))
    lua = replace_lua_variable(lua, 'ShipbaseSize',    to_num(row['Size']))
    lua = replace_lua_variable(lua, 'Shipcost',        to_num(row['Points']))
    lua = replace_lua_variable(lua, 'ShipScan',        to_num(row['Scan']))
    lua = replace_lua_variable(lua, 'ShipThrust',      to_num(row['Thrust']))
    lua = replace_lua_variable(lua, 'Signature',       to_num(row['Sig']))
    lua = replace_lua_variable(lua, 'SHIPCardURL',     row['Card Image'],  is_string=True)
    lua = replace_lua_variable(lua, 'FIXED_IMAGE_URL', row['Model Image'], is_string=True)
    lua = replace_lua_variable(lua, 'Model3DExists',   str(model_3d_exists).lower())
    lua = replace_lua_variable(lua, 'Vectored',        str(is_vectored).lower())
    lua = replace_lua_variable(lua, 'Monitor',         str(is_monitored).lower())

    if model_3d_exists:
        lua = replace_lua_variable(lua, 'TARGET_MESH_URL',    model_file,   is_string=True)
        lua = replace_lua_variable(lua, 'TARGET_DIFFUSE_URL', texture_file, is_string=True)

    # Patch: allow ShipThrust=0 for stations. The template has a safety check
    # that treats 0 as invalid and overrides to 9. Fix it to only default on nil.
    lua = lua.replace(
        'if not ShipThrust or ShipThrust == 0 then\r\n        ShipThrust = 9  -- Default value',
        'if not ShipThrust then\r\n        ShipThrust = 9  -- Default value'
    )
    lua = lua.replace(
        'if not ShipThrust or ShipThrust == 0 then\n        ShipThrust = 9  -- Default value',
        'if not ShipThrust then\n        ShipThrust = 9  -- Default value'
    )

    ship['LuaScript'] = lua

    ship['LuaScriptState'] = update_lua_script_state(
        ship.get('LuaScriptState', ''), row
    )

    for child in ship.get('ChildObjects', []):
        if child.get('Nickname') == 'SHIP3dmodel':
            if model_3d_exists:
                child['CustomMesh']['MeshURL']    = model_file
                child['CustomMesh']['DiffuseURL'] = texture_file
            break

    return ship


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Usage: python tts_ship_generator.py <tts_save.json> <ship_data.xlsx> [output.json] [template_save.json]")
        print()
        print("  tts_save.json       - Save file with faction container tiles")
        print("  ship_data.xlsx      - Excel spreadsheet with ship data")
        print("  output.json         - Output file (default: TTS_Output.json)")
        print("  template_save.json  - Optional: separate save containing the template ship model")
        sys.exit(1)

    save_path     = sys.argv[1]
    xlsx_path     = sys.argv[2]
    output_path   = sys.argv[3] if len(sys.argv) > 3 else "TTS_Output.json"
    template_path = sys.argv[4] if len(sys.argv) > 4 else None

    # 1. Load data
    print(f"Loading TTS save: {save_path}")
    with open(save_path, 'r', encoding='utf-8') as f:
        save_data = json.load(f)

    print(f"Loading ship data: {xlsx_path}")
    rows = read_ship_rows(xlsx_path)
    print(f"  Found {len(rows)} ships in spreadsheet.")

    # 2. Find the template object
    template_idx = None
    template_obj = None

    for i, obj in enumerate(save_data.get('ObjectStates', [])):
        if 'SHIP_ID' in obj.get('LuaScript', ''):
            template_idx = i
            template_obj = obj
            print(f"  Template found in main save: index {i}, "
                  f"Nickname='{obj.get('Nickname')}', GUID={obj['GUID']}")
            break

    if template_obj is None:
        if template_path is None:
            print("ERROR: No template object found. Provide a template save as the 4th argument.")
            sys.exit(1)
        print(f"  Loading template from: {template_path}")
        with open(template_path, 'r', encoding='utf-8') as f:
            template_data = json.load(f)
        _, template_obj = find_template_object(template_data)
        print(f"  Template found: Nickname='{template_obj.get('Nickname')}', "
              f"GUID={template_obj['GUID']}")

    # 3. Discover faction container tiles
    faction_tiles = {}
    for i, obj in enumerate(save_data.get('ObjectStates', [])):
        nick = obj.get('Nickname', '')
        if nick.startswith('Faction '):
            pdf_id = nick.split('Faction ', 1)[1].strip()
            faction_tiles[pdf_id] = i
            print(f"  Found container tile: '{nick}' (PDF ID '{pdf_id}') at index {i}")

    if not faction_tiles:
        print("ERROR: No faction container tiles found (expected 'Faction BIO', 'Faction UCM', etc.)")
        sys.exit(1)

    # 4. Collect existing GUIDs
    existing_guids = collect_all_guids(save_data)

    # 5. Process each ship row
    groups: dict[str, list] = {}
    vectored_count  = 0
    monitored_count = 0

    for i, row in enumerate(rows):
        pdf_id    = str(row.get('PDF ID', 'UNKNOWN')).strip()
        ship_name = str(row['Name'])
        ship_type = str(row.get('Type', ''))
        card_url  = str(row.get('Card Image', ''))
        base_size = to_num(row['Size'])

        # Track special flags before creating ship
        is_vectored, is_monitored = parse_special_flags(row)
        if is_vectored:
            vectored_count += 1
        if is_monitored:
            monitored_count += 1

        # a) Create the ship model from template
        ship = create_ship_from_template(template_obj, row, existing_guids)

        # b) Wrap it as a spawner tile
        tile_guid = generate_guid(existing_guids)
        tile = make_spawner_tile(ship, card_url, tile_guid)

        # c) Serialize the tile to JSON
        tile_json = json.dumps(tile, ensure_ascii=False)

        # d) Build the card entry
        card_entry = {
            "name":        ship_name,
            "imageURL":    card_url,
            "tonnage":     detect_class_tonnage(ship_name, ship_type),
            "shipTonnage": detect_ship_tonnage(base_size),
            "baseScale":   1,
            "json":        tile_json,
        }

        groups.setdefault(pdf_id, []).append(card_entry)

        if (i + 1) % 50 == 0:
            print(f"  Processed {i + 1}/{len(rows)} ships...")

    total_ships = len(rows)
    print(f"  Created {total_ships} ship entries across {len(groups)} factions.")
    print(f"  Special flags detected:")
    print(f"    isVectored  = true : {vectored_count} ships")
    print(f"    isMonitored = true : {monitored_count} ships")
    print(f"    Neither flag       : {total_ships - vectored_count - monitored_count} ships")

    # 6. Sort each group alphabetically
    for pdf_id in groups:
        groups[pdf_id].sort(key=lambda c: c['name'].lower())

    # 7. Embed card data into each faction container tile's Lua script
    for pdf_id, card_entries in groups.items():
        if pdf_id not in faction_tiles:
            print(f"  WARNING: No container tile found for PDF ID '{pdf_id}' "
                  f"({len(card_entries)} ships orphaned)")
            continue

        tile_idx  = faction_tiles[pdf_id]
        container = save_data['ObjectStates'][tile_idx]
        lua_script = container['LuaScript']

        # Remove any existing embedded data section (use last occurrence)
        for marker in ("\r\n--EMBEDDED_DATA_START", "\n--EMBEDDED_DATA_START"):
            pos = lua_script.rfind(marker)
            if pos != -1:
                lua_script = lua_script[:pos]
                break

        embedded   = build_embedded_data(card_entries)
        lua_script = lua_script + embedded

        container['LuaScript'] = lua_script
        print(f"  Embedded {len(card_entries)} ships into '{container['Nickname']}'")

    # 8. Remove the template object from the save (only if it was in the main save)
    if template_idx is not None:
        save_data['ObjectStates'] = [
            obj for i, obj in enumerate(save_data['ObjectStates'])
            if i != template_idx
        ]

    # 9. Write output
    print(f"Writing output to: {output_path}")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(save_data, f, ensure_ascii=False)

    print("Done!")
    print(f"  Total objects on table: {len(save_data['ObjectStates'])}")


if __name__ == '__main__':
    main()
