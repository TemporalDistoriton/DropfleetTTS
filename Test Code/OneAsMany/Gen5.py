#!/usr/bin/env python3
"""
TTS Ship Generator v5 — Template-Optimised (with Spawner Tiles)
================================================================
Key optimisation over v3/v4: instead of storing the full ~190KB spawner-tile
JSON per card entry (which itself contains the ~170KB ship model with ~146KB
of duplicated LuaScript), this version:

  1. Extracts the ship LuaScript into two parts:
       - Per-ship variable header (~1KB, unique per ship)
       - Template body (~145KB, identical for all ships)
  2. Stores THREE templates ONCE per faction container:
       - shipScriptTemplate   — the shared ship Lua body (~145KB)
       - spawnerScriptTemplate — the spawner-tile Lua (~1.2KB)
       - spawnerTileTemplate  — the spawner-tile TTS JSON shell (~800B)
  3. Each card stores only:
       - shipVars    — compact Lua table of per-ship variables (~600B)
       - shipObjJson — ship model JSON with scripts stripped (~13KB)
       - shipState   — the LuaScriptState JSON string (~2.5KB)
  4. At spawn time (when player picks ships in the UI), the faction container:
       a) Reconstructs the full ship LuaScript from shipVars + template
       b) Injects it into the stripped ship JSON via JSON.decode/encode
       c) Wraps the ship JSON inside a spawner-tile Lua as objectJSON
       d) Builds the spawner tile TTS JSON and spawns it on the sideboard
     The spawner tile on the sideboard is identical to what v4 produces.

Result: ~89% reduction in embedded data size per faction container.
  e.g. UCM with 67 ships: 12.8MB -> 1.4MB

Backward compatible: cards with legacy `json` field still spawn as before.

Usage:
    python Gen5.py <tts_save.json> <ship_data.xlsx> [output.json] [template_save.json]
"""

import json
import copy
import re
import sys
import random
import openpyxl


# --- Helpers ----------------------------------------------------------------

def generate_guid(existing_guids: set) -> str:
    """Generate a unique 6-char hex GUID not already in use."""
    while True:
        guid = ''.join(random.choices('0123456789abcdef', k=6))
        if guid not in existing_guids:
            existing_guids.add(guid)
            return guid


def collect_all_guids(save_data: dict) -> set:
    """Walk the entire save and collect every GUID already in use."""
    guids = set()
    def _walk(obj):
        if isinstance(obj, dict):
            if 'GUID' in obj:
                guids.add(obj['GUID'])
            for v in obj.values():
                _walk(v)
        elif isinstance(obj, list):
            for item in obj:
                _walk(item)
    _walk(save_data)
    return guids


def to_num(val):
    """Convert a value to int or float; pass through if already numeric."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val if not isinstance(val, float) or val != int(val) else int(val)
    try:
        s = str(val).strip()
        if '.' in s:
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        return 0


def parse_special_flags(row: dict) -> tuple:
    """Parse Special column for isVectored and isMonitored flags."""
    special = str(row.get('Special', '') or '').lower()
    is_vectored  = 'vectored' in special
    is_monitored = 'monitor'  in special
    return is_vectored, is_monitored


# --- Tonnage classification -------------------------------------------------

TONNAGE_PATTERNS = [
    ("dreadnought",     "Dreadnaught"),
    ("dreadnaught",     "Dreadnaught"),
    ("super battleship","Super Battleship"),
    ("superbattleship", "Super Battleship"),
    ("supercarrier",    "Supercarrier"),
    ("battleship",      "Battleship"),
    ("battlecruiser",   "Battlecruiser"),
    ("troopship",       "Troopship"),
    ("heavy cruiser",   "Heavy Cruiser"),
    ("light cruiser",   "Light Cruiser"),
    ("cruiser",         "Cruiser"),
    ("runner",          "Runner"),
    ("destroyer",       "Destroyer"),
    ("cutter",          "Cutter"),
    ("monitor",         "Monitor"),
    ("carrier",         "Carrier"),
    ("frigate",         "Frigate"),
    ("lighter",         "Lighter"),
    ("corvette",        "Corvette"),
    ("cell",            "Cell"),
]

SHIPBASE_TO_TONNAGE = {
    20: "Light",
    30: "Light",
    40: "Medium",
    50: "Heavy",
    60: "Colossal",
}


def detect_class_tonnage(name: str, description: str) -> str:
    text = (description.lower() + " " + name.lower())
    for pattern, tonnage in TONNAGE_PATTERNS:
        if pattern in text:
            return tonnage
    return "Other"


def detect_ship_tonnage(base_size: int) -> str:
    return SHIPBASE_TO_TONNAGE.get(base_size, "Light")


# --- Template extraction ----------------------------------------------------

TEMPLATE_SPLIT_MARKER = "local playerColor"


def extract_script_template(lua: str) -> str:
    """
    Extract the template body from the ship LuaScript.
    Returns everything from 'local playerColor' onward — the shared code
    that is identical for every ship.
    """
    idx = lua.find(TEMPLATE_SPLIT_MARKER)
    if idx == -1:
        raise RuntimeError(
            f"Could not find '{TEMPLATE_SPLIT_MARKER}' split point in template ship script. "
            "Is the ship script V8.2 or compatible?"
        )
    body = lua[idx:]

    # Apply the ShipThrust=0 station fix to the template once
    for nl in ('\r\n', '\n'):
        body = body.replace(
            f'if not ShipThrust or ShipThrust == 0 then{nl}        ShipThrust = 9  -- Default value',
            f'if not ShipThrust then{nl}        ShipThrust = 9  -- Default value'
        )

    return body


# --- Spawner tile templates (Python constants) ------------------------------

# The spawner-tile Lua script WITHOUT the objectJSON assignment.
# Uses self.getName() instead of a hardcoded ship name so the template is
# reusable across all ships.
SPAWNER_LUA_TEMPLATE = """\
-- SPAWNER TILE
-- Click the button to spawn the model on this tile

function onLoad()
  local pos = self.getPosition()
  rebuildUI()
end

function rebuildUI()
    local ui = {
        {tag='Defaults', children={
            {tag='Text', attributes={color='#cccccc', fontSize='18', alignment='MiddleLeft'}},
            {tag='InputField', attributes={fontSize='24', preferredHeight='40'}},
            {tag='ToggleButton', attributes={fontSize='18', preferredHeight='40', colors='#ffcc33|#ffffff|#808080|#606060', selectedBackgroundColor='#dddddd', deselectedBackgroundColor='#999999'}},
            {tag='Button', attributes={fontSize='12',textColor='#111111', preferredHeight='40', colors='#dddddd|#ffffff|#808080|#f6f6f6'}},
            {tag='Toggle', attributes={textColor='#cccccc'}},
        }},

        {tag='button', attributes={onClick='ui_createModel',text='Spawn Ship',  colors='#ccccccff|#ffffffff|#404040ff|#808080ff', width='120', height='20', position='0 110 -5', rotation='0 0 180' }}
    }

    self.UI.setXmlTable(ui)
end

function ui_createModel(player, value, id)
    local color = player.color
    local pos = self.getPosition()
    local rot = {x = 0, y = 0, z = 0}
    if color == "Blue" then
        rot.y = 180
    end
    spawnObjectJSON({
        json = objectJSON,
        position = {x = pos.x, y = pos.y + 2, z = pos.z},
        rotation = rot,
        callback_function = function(spawned_object)
            Wait.frames(function()
                spawned_object.call("SetPlayerColor", {playerColor = color})
            end, 30)
        end
    })
    broadcastToAll("Spawned: " .. self.getName() .. " (" .. color .. ")", {0, 1, 0})
end
"""

# The spawner-tile TTS object JSON shell — everything except Nickname,
# Description, CustomImage URLs, and LuaScript (filled in at spawn time).
SPAWNER_TILE_JSON_TEMPLATE = {
    "Name": "Custom_Tile",
    "Transform": {
        "posX": 0, "posY": 1.0, "posZ": 0,
        "rotX": 0, "rotY": 0, "rotZ": 0,
        "scaleX": 2.0, "scaleY": 1.0, "scaleZ": 2.0,
    },
    "Nickname": "",
    "Description": "",
    "GMNotes": "",
    "AltLookAngle": {"x": 0.0, "y": 0.0, "z": 0.0},
    "ColorDiffuse": {"r": 1.0, "g": 1.0, "b": 1.0},
    "LayoutGroupSortIndex": 0,
    "Value": 0,
    "Locked": False,
    "Grid": True,
    "Snap": True,
    "IgnoreFoW": False,
    "MeasureMovement": False,
    "DragSelectable": True,
    "Autoraise": True,
    "Sticky": True,
    "Tooltip": True,
    "GridProjection": False,
    "HideWhenFaceDown": False,
    "Hands": False,
    "CustomImage": {
        "ImageURL": "",
        "ImageSecondaryURL": "",
        "ImageScalar": 1.0,
        "WidthScale": 0.0,
        "CustomTile": {
            "Type": 0,
            "Thickness": 0.1,
            "Stackable": False,
            "Stretch": True,
        },
    },
    "LuaScript": "",
    "LuaScriptState": "",
    "XmlUI": "",
}


# --- Per-ship data builders -------------------------------------------------

def build_ship_vars(row: dict) -> dict:
    """Build the compact per-ship variables dict from one Excel row."""
    model_file   = str(row.get('3D MODEL FILE', 'None')).strip()
    texture_file = str(row.get('3D MODEL TEXTURE', 'None')).strip()
    model_3d_exists = model_file not in ('None', '', 'none')
    is_vectored, is_monitored = parse_special_flags(row)

    return {
        'SHIP_ID':           str(row['Name']),
        'Shiphealth':        to_num(row['Hull']),
        'ShipbaseSize':      to_num(row['Size']),
        'Shipcost':          to_num(row['Points']),
        'ShipScan':          to_num(row['Scan']),
        'ShipThrust':        to_num(row['Thrust']),
        'Signature':         to_num(row['Sig']),
        'SHIPCardURL':       str(row['Card Image']),
        'FIXED_IMAGE_URL':   str(row['Model Image']),
        'Model3DExists':     model_3d_exists,
        'TARGET_MESH_URL':   model_file if model_3d_exists else '',
        'TARGET_DIFFUSE_URL':texture_file if model_3d_exists else '',
        'Vectored':          is_vectored,
        'Monitor':           is_monitored,
    }


def build_lua_script_state(template_state_json: str, row: dict) -> str:
    """
    Build the LuaScriptState JSON for one ship, starting from the template's
    state and patching in per-ship values.
    """
    if not template_state_json or not template_state_json.strip():
        return ''
    try:
        state = json.loads(template_state_json)
    except json.JSONDecodeError:
        return template_state_json

    is_vectored, is_monitored = parse_special_flags(row)

    state['shipID']          = row['Name']
    state['Shiphealth']      = to_num(row['Hull'])
    state['ShipbaseSize']    = to_num(row['Size'])
    state['Shipcost']        = to_num(row['Points'])
    state['ShipScan']        = to_num(row['Scan'])
    state['ShipThrust']      = to_num(row['Thrust'])
    state['Signature']       = to_num(row['Sig'])
    state['SHIPCardURL']     = row['Card Image']
    state['FIXED_IMAGE_URL'] = row['Model Image']
    state['baseSignature']   = to_num(row['Sig'])
    state['Vectored']        = is_vectored
    state['Monitor']         = is_monitored

    model_3d_exists = str(row.get('3D MODEL FILE', 'None')).strip() not in ('None', '', 'none')
    state['Model3DAvailable'] = model_3d_exists

    for block_key in ('originalData', 'state'):
        block = state.get(block_key)
        if isinstance(block, dict):
            if 'health' in block:
                hp = to_num(row['Hull'])
                block['health']['current'] = hp
                block['health']['max']     = hp
            if 'base' in block:
                block['base']['size'] = to_num(row['Size'])
            if 'sig' in block:
                block['sig'] = to_num(row['Sig'])

    return json.dumps(state)


def build_stripped_ship_obj(template: dict, row: dict, existing_guids: set) -> tuple:
    """
    Create a ship model object with LuaScript="" and LuaScriptState="".
    Returns (stripped_json_str, lua_script_state_str).
    """
    ship = copy.deepcopy(template)

    # Unique GUIDs
    ship['GUID'] = generate_guid(existing_guids)
    for child in ship.get('ChildObjects', []):
        child['GUID'] = generate_guid(existing_guids)

    # Per-ship metadata
    ship['Nickname']    = str(row['Name'])
    ship['Description'] = str(row.get('Type', ''))

    # Update 3D model child
    model_file   = str(row.get('3D MODEL FILE', 'None')).strip()
    texture_file = str(row.get('3D MODEL TEXTURE', 'None')).strip()
    model_3d_exists = model_file not in ('None', '', 'none')

    for child in ship.get('ChildObjects', []):
        if child.get('Nickname') == 'SHIP3dmodel':
            if model_3d_exists:
                child['CustomMesh']['MeshURL']    = model_file
                child['CustomMesh']['DiffuseURL'] = texture_file
            break

    # Build state before stripping
    ship_state = build_lua_script_state(ship.get('LuaScriptState', ''), row)

    # Strip scripts - these get reconstructed at spawn time
    ship['LuaScript']      = ''
    ship['LuaScriptState'] = ''

    return json.dumps(ship, ensure_ascii=False), ship_state


# --- Lua long-string delimiter safety --------------------------------------

def pick_safe_delimiter(text: str, start_level: int = 4) -> str:
    """Find a [=...=[ delimiter level that doesn't appear in the text."""
    level = '=' * start_level
    while f']={level}]' in text:
        level += '='
    return level


# --- Lua injection: functions added to the embedded data section ------------

LUA_HELPER_FUNCTIONS = r"""
--==============================================================================
-- TEMPLATE RECONSTRUCTION ENGINE (auto-generated by Gen5.py)
--==============================================================================

--- Build the per-ship variable declaration header from a shipVars table.
function buildScriptHeader(vars)
    if not vars then return "" end
    local lines = {}
    table.insert(lines, "--- SHIP SCRIPT V8.2 (Template-Reconstructed)")
    table.insert(lines, "-------------------------------------------------------------------------------")
    table.insert(lines, "-- Variables That Change Per ship")
    table.insert(lines, 'local SHIP_ID = "' .. tostring(vars.SHIP_ID or "Unknown") .. '"')
    table.insert(lines, "local Shiphealth = " .. tostring(vars.Shiphealth or 2))
    table.insert(lines, "local ShipbaseSize = " .. tostring(vars.ShipbaseSize or 30))
    table.insert(lines, "local Shipcost = " .. tostring(vars.Shipcost or 2))
    table.insert(lines, "local ShipScan = " .. tostring(vars.ShipScan or 2))
    table.insert(lines, "local ShipThrust = " .. tostring(vars.ShipThrust or 2))
    table.insert(lines, "local Signature = " .. tostring(vars.Signature or 2))
    table.insert(lines, 'local SHIPCardURL = "' .. tostring(vars.SHIPCardURL or "") .. '"')
    table.insert(lines, 'local FIXED_IMAGE_URL = "' .. tostring(vars.FIXED_IMAGE_URL or "") .. '"')
    table.insert(lines, "local Model3DExists = " .. tostring(vars.Model3DExists or false))
    table.insert(lines, 'local TARGET_MESH_URL = "' .. tostring(vars.TARGET_MESH_URL or "") .. '"')
    table.insert(lines, 'local TARGET_DIFFUSE_URL = "' .. tostring(vars.TARGET_DIFFUSE_URL or "") .. '"')
    table.insert(lines, "")
    table.insert(lines, "-- Ship-type flags")
    table.insert(lines, "local Vectored = " .. tostring(vars.Vectored or false))
    table.insert(lines, "local Monitor  = " .. tostring(vars.Monitor or false))
    table.insert(lines, "")
    table.insert(lines, "")
    return table.concat(lines, "\n")
end

--- Reconstruct the full ship JSON from template + per-ship data.
--- Uses TTS's built-in JSON module to safely inject the LuaScript.
function buildFullShipJSON(card)
    if not card or not card.shipVars or not card.shipObjJson then return nil end
    if not shipScriptTemplate or shipScriptTemplate == "" then return nil end

    local fullScript = buildScriptHeader(card.shipVars) .. shipScriptTemplate
    local shipObj = JSON.decode(card.shipObjJson)
    if not shipObj then
        broadcastToAll("ERROR: Failed to decode ship JSON for " .. (card.name or "?"), {1, 0, 0})
        return nil
    end

    shipObj.LuaScript      = fullScript
    shipObj.LuaScriptState = card.shipState or ""

    return JSON.encode(shipObj)
end

--- Find a safe Lua long-string delimiter level for the given text.
--- Returns the equals string (e.g. "======" for [======[...]======]).
function findSafeDelimiter(text)
    local level = 6
    while true do
        local closeDelim = "]" .. string.rep("=", level) .. "]"
        if not text:find(closeDelim, 1, true) then
            return string.rep("=", level)
        end
        level = level + 1
    end
end

--- Build the full spawner-tile JSON string from a card entry.
--- Reconstructs the ship -> wraps it in a spawner tile -> returns spawn-ready JSON.
function buildSpawnerTileJSON(card)
    -- 1. Reconstruct the full ship JSON string
    local shipJson = buildFullShipJSON(card)
    if not shipJson then return nil end

    -- 2. Build the spawner-tile Lua: script template + objectJSON assignment
    local eq = findSafeDelimiter(shipJson)
    local spawnerLua = spawnerScriptTemplate
        .. "\nobjectJSON = [" .. eq .. "[" .. shipJson .. "]" .. eq .. "]"

    -- 3. Clone the spawner tile JSON template and fill in per-ship values
    local tile = JSON.decode(spawnerTileTemplate)
    if not tile then
        broadcastToAll("ERROR: Failed to decode spawner tile template", {1, 0, 0})
        return nil
    end

    tile.Nickname    = card.name or "Unknown"
    tile.Description = card.description or ""
    tile.LuaScript   = spawnerLua

    if tile.CustomImage then
        tile.CustomImage.ImageURL          = card.imageURL or ""
        tile.CustomImage.ImageSecondaryURL = card.imageURL or ""
    end

    return JSON.encode(tile)
end

--==============================================================================
-- OVERRIDES: spawnCardsForPlayer (handles both new template + legacy formats)
--==============================================================================

function spawnCardsForPlayer(playerColor, cardIndices)
    if not playerSpawnCounts then playerSpawnCounts = {} end

    local spawnConfig = CONFIG.PLAYER_SPAWN_POSITIONS[playerColor]
    if not spawnConfig then
        spawnConfig = {
            position  = {x = 0, y = 1, z = 0},
            rotation  = {x = 0, y = 0, z = 0},
            direction = {x = 1, y = 0, z = 0},
        }
        broadcastToColor(
            "Warning: No spawn position configured for " .. playerColor .. ", using center.",
            playerColor, {1, 1, 0}
        )
    end

    local startPos     = spawnConfig.position
    local rotation     = spawnConfig.rotation
    local direction    = spawnConfig.direction
    local spacing      = CONFIG.CARD_SPAWN_SPACING
    local cardsPerRow  = CONFIG.CARDS_PER_ROW or 6

    local rowDirection = {
        x = direction.z,
        y = 0,
        z = direction.x,
    }

    local currentCount = playerSpawnCounts[playerColor] or 0

    for i, cardIndex in ipairs(cardIndices) do
        local card = savedShipCards[cardIndex]
        if card then
            local spawnJson = nil

            -- New template format: reconstruct spawner tile on-the-fly
            if card.shipObjJson and card.shipVars then
                spawnJson = buildSpawnerTileJSON(card)
            -- Legacy format: pre-built spawner tile JSON
            elseif card.json then
                spawnJson = card.json
            end

            if spawnJson then
                local col = currentCount % cardsPerRow
                local row = math.floor(currentCount / cardsPerRow)
                local colOffset = col * spacing
                local rowOffset = row * spacing
                local spawnPos = {
                    x = startPos.x + (direction.x * colOffset) + (rowDirection.x * rowOffset),
                    y = startPos.y,
                    z = startPos.z + (direction.z * colOffset) + (rowDirection.z * rowOffset),
                }

                spawnObjectJSON({
                    json     = spawnJson,
                    position = spawnPos,
                    rotation = rotation,
                    sound    = true,
                })

                currentCount = currentCount + 1
            end
        end
    end

    playerSpawnCounts[playerColor] = currentCount
end

--==============================================================================
-- OVERRIDE: onLoad (handles both new template fields + legacy stripping)
--==============================================================================

function onLoad(saved_data)
    selfGUID = self.getGUID()

    if savedShipCards == nil then savedShipCards = {} end
    if fresh == nil then fresh = true end
    if #savedShipCards > 0 then fresh = false end

    -- Initialise player state tables
    playerSelections       = {}
    playerCardIndices      = {}
    playerSearchFilters    = {}
    playerActiveTabs       = {}
    playerActiveTonnageTabs = {}
    playerSpawnCounts      = {}

    -- Clean loaded data (strip long-string delimiters if present)
    for _, card in ipairs(savedShipCards) do
        -- Legacy format
        if card.json        then card.json        = stripBrackets(card.json) end
        if card.imageURL    then card.imageURL    = stripBrackets(card.imageURL) end
        -- New template format
        if card.shipObjJson then card.shipObjJson = stripBrackets(card.shipObjJson) end
        if card.shipState   then card.shipState   = stripBrackets(card.shipState) end
        if card.description then card.description = stripBrackets(card.description) end
        if card.shipVars then
            for k, v in pairs(card.shipVars) do
                if type(v) == "string" then
                    card.shipVars[k] = stripBrackets(v)
                end
            end
        end
    end

    -- Strip templates too (safety)
    if shipScriptTemplate    then shipScriptTemplate    = stripBrackets(shipScriptTemplate) end
    if spawnerScriptTemplate then spawnerScriptTemplate = stripBrackets(spawnerScriptTemplate) end
    if spawnerTileTemplate   then spawnerTileTemplate   = stripBrackets(spawnerTileTemplate) end

    ensureShipTonnage()

    if fresh then
        createSetupButton()
    else
        createMainButtonsWithList()
    end
end

--==============================================================================
-- OVERRIDE: updateSave (writes new template format + legacy compat)
--==============================================================================

function updateSave()
    local script = self.getLuaScript()

    local dataStartPattern = "\n%-%-EMBEDDED_DATA_START"
    local dataStart = script:find(dataStartPattern)
    if dataStart then
        script = script:sub(1, dataStart - 1)
    end

    local embeddedData = "\n--EMBEDDED_DATA_START\n"

    -- Write the ship script template (if available)
    if shipScriptTemplate and shipScriptTemplate ~= "" then
        embeddedData = embeddedData .. "shipScriptTemplate = [====[" .. shipScriptTemplate .. "]====]\n\n"
    end

    -- Write the spawner templates (if available)
    if spawnerScriptTemplate and spawnerScriptTemplate ~= "" then
        embeddedData = embeddedData .. "spawnerScriptTemplate = [====[" .. spawnerScriptTemplate .. "]====]\n\n"
    end

    if spawnerTileTemplate and spawnerTileTemplate ~= "" then
        embeddedData = embeddedData .. "spawnerTileTemplate = [====[" .. spawnerTileTemplate .. "]====]\n\n"
    end

    embeddedData = embeddedData .. "savedShipCards = {\n"

    for i, card in ipairs(savedShipCards) do
        embeddedData = embeddedData .. "  {\n"
        embeddedData = embeddedData .. "    name = [====[" .. (card.name or "Unknown") .. "]====],\n"
        embeddedData = embeddedData .. "    imageURL = [====[" .. (card.imageURL or "") .. "]====],\n"
        embeddedData = embeddedData .. "    tonnage = [====[" .. (card.tonnage or "Light") .. "]====],\n"
        embeddedData = embeddedData .. "    shipTonnage = [====[" .. (card.shipTonnage or "Light") .. "]====],\n"
        embeddedData = embeddedData .. "    baseScale = " .. (card.baseScale or 1) .. ",\n"

        if card.shipVars and card.shipObjJson then
            -- New template format
            embeddedData = embeddedData .. "    description = [====[" .. (card.description or "") .. "]====],\n"
            embeddedData = embeddedData .. "    shipVars = {\n"
            for _, key in ipairs({"SHIP_ID", "SHIPCardURL", "FIXED_IMAGE_URL", "TARGET_MESH_URL", "TARGET_DIFFUSE_URL"}) do
                local v = card.shipVars[key]
                if v then
                    embeddedData = embeddedData .. "      " .. key .. " = [====[" .. tostring(v) .. "]====],\n"
                end
            end
            for _, key in ipairs({"Shiphealth", "ShipbaseSize", "Shipcost", "ShipScan", "ShipThrust", "Signature"}) do
                local v = card.shipVars[key]
                if v ~= nil then
                    embeddedData = embeddedData .. "      " .. key .. " = " .. tostring(v) .. ",\n"
                end
            end
            for _, key in ipairs({"Model3DExists", "Vectored", "Monitor"}) do
                local v = card.shipVars[key]
                if v ~= nil then
                    embeddedData = embeddedData .. "      " .. key .. " = " .. tostring(v) .. ",\n"
                end
            end
            embeddedData = embeddedData .. "    },\n"
            embeddedData = embeddedData .. "    shipState = [====[" .. (card.shipState or "") .. "]====],\n"
            embeddedData = embeddedData .. "    shipObjJson = [====[" .. (card.shipObjJson or "") .. "]====],\n"
        elseif card.json then
            -- Legacy format
            embeddedData = embeddedData .. "    json = [====[" .. card.json .. "]====],\n"
        end

        embeddedData = embeddedData .. "  },\n"
    end

    embeddedData = embeddedData .. "}\n\n"
    embeddedData = embeddedData .. "fresh = " .. tostring(fresh) .. "\n"

    script = script .. embeddedData
    self.setLuaScript(script)
end
"""


# --- Embedded data generation -----------------------------------------------

def build_embedded_data(card_entries: list, script_template: str) -> str:
    """
    Build the embedded Lua data section with:
      - Helper Lua functions (override spawnCardsForPlayer, onLoad, updateSave)
      - Ship script template (stored once)
      - Spawner tile templates (stored once)
      - Compact card entries (shipVars + stripped JSON per ship)
    """
    lines = []
    lines.append("\n--EMBEDDED_DATA_START")

    # 1. Lua helper functions and overrides
    lines.append(LUA_HELPER_FUNCTIONS)

    # 2. Ship script template (stored once, shared by all cards)
    delim = pick_safe_delimiter(script_template)
    lines.append(f"shipScriptTemplate = [={delim}[{script_template}]={delim}]")
    lines.append("")

    # 3. Spawner script template (stored once)
    sp_delim = pick_safe_delimiter(SPAWNER_LUA_TEMPLATE)
    lines.append(f"spawnerScriptTemplate = [={sp_delim}[{SPAWNER_LUA_TEMPLATE}]={sp_delim}]")
    lines.append("")

    # 4. Spawner tile JSON template (stored once)
    tile_json = json.dumps(SPAWNER_TILE_JSON_TEMPLATE, ensure_ascii=False)
    tile_delim = pick_safe_delimiter(tile_json)
    lines.append(f"spawnerTileTemplate = [={tile_delim}[{tile_json}]={tile_delim}]")
    lines.append("")

    # 5. Card data
    lines.append("savedShipCards = {")

    for card in card_entries:
        lines.append("  {")
        lines.append(f"    name = [====[{card['name']}]====],")
        lines.append(f"    description = [====[{card['description']}]====],")
        lines.append(f"    imageURL = [====[{card['imageURL']}]====],")
        lines.append(f"    tonnage = [====[{card['tonnage']}]====],")
        lines.append(f"    shipTonnage = [====[{card['shipTonnage']}]====],")
        lines.append(f"    baseScale = {card['baseScale']},")

        # Per-ship variables (compact Lua table)
        v = card['shipVars']
        lines.append("    shipVars = {")
        for key in ('SHIP_ID', 'SHIPCardURL', 'FIXED_IMAGE_URL', 'TARGET_MESH_URL', 'TARGET_DIFFUSE_URL'):
            lines.append(f"      {key} = [====[{v.get(key, '')}]====],")
        for key in ('Shiphealth', 'ShipbaseSize', 'Shipcost', 'ShipScan', 'ShipThrust', 'Signature'):
            lines.append(f"      {key} = {v.get(key, 0)},")
        for key in ('Model3DExists', 'Vectored', 'Monitor'):
            lines.append(f"      {key} = {str(v.get(key, False)).lower()},")
        lines.append("    },")

        # LuaScriptState and stripped ship JSON
        lines.append(f"    shipState = [====[{card['shipState']}]====],")
        lines.append(f"    shipObjJson = [====[{card['shipObjJson']}]====],")
        lines.append("  },")

    lines.append("}")
    lines.append("")
    lines.append("fresh = false")
    return "\n".join(lines)


# --- Ship template discovery ------------------------------------------------

def find_template_object(save_data: dict) -> tuple:
    for i, obj in enumerate(save_data.get('ObjectStates', [])):
        if 'SHIP_ID' in obj.get('LuaScript', ''):
            return i, obj
    raise RuntimeError("No template object with SHIP_ID found in the save file.")


def read_ship_rows(xlsx_path: str) -> list:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if row_dict.get('Name') is None:
            continue
        rows.append(row_dict)
    return rows


# --- Main -------------------------------------------------------------------

def main():
    if len(sys.argv) < 3:
        print("Usage: python Gen5.py <tts_save.json> <ship_data.xlsx> [output.json] [template_save.json]")
        print()
        print("  tts_save.json       - Save file with faction container tiles (e.g. MT.json)")
        print("  ship_data.xlsx      - Excel spreadsheet with ship data")
        print("  output.json         - Output file (default: TTS_Output.json)")
        print("  template_save.json  - Optional: separate save containing the template ship model")
        sys.exit(1)

    save_path     = sys.argv[1]
    xlsx_path     = sys.argv[2]
    output_path   = sys.argv[3] if len(sys.argv) > 3 else "TTS_Output.json"
    template_path = sys.argv[4] if len(sys.argv) > 4 else None

    # -- 1. Load data --
    print(f"Loading TTS save: {save_path}")
    with open(save_path, 'r', encoding='utf-8') as f:
        save_data = json.load(f)

    print(f"Loading ship data: {xlsx_path}")
    rows = read_ship_rows(xlsx_path)
    print(f"  Found {len(rows)} ships in spreadsheet.")

    # -- 2. Find the template object --
    template_idx = None
    template_obj = None

    for i, obj in enumerate(save_data.get('ObjectStates', [])):
        if 'SHIP_ID' in obj.get('LuaScript', ''):
            template_idx = i
            template_obj = obj
            print(f"  Template found in main save: index {i}, "
                  f"Nickname='{obj.get('Nickname')}', GUID={obj['GUID']}")
            break

    if template_obj is None:
        if template_path is None:
            print("ERROR: No template object found. Provide a template save as the 4th argument.")
            sys.exit(1)
        print(f"  Loading template from: {template_path}")
        with open(template_path, 'r', encoding='utf-8') as f:
            template_data = json.load(f)
        _, template_obj = find_template_object(template_data)
        print(f"  Template found: Nickname='{template_obj.get('Nickname')}', "
              f"GUID={template_obj['GUID']}")

    # -- 3. Extract the shared script template --
    template_lua = template_obj['LuaScript']
    script_template = extract_script_template(template_lua)
    print(f"  Script template extracted: {len(script_template):,} chars "
          f"(shared across all ships)")

    # -- 4. Discover faction container tiles --
    faction_tiles = {}
    for i, obj in enumerate(save_data.get('ObjectStates', [])):
        nick = obj.get('Nickname', '')
        if nick.startswith('Faction '):
            pdf_id = nick.split('Faction ', 1)[1].strip()
            faction_tiles[pdf_id] = i
            print(f"  Found container tile: '{nick}' (PDF ID '{pdf_id}') at index {i}")

    if not faction_tiles:
        print("ERROR: No faction container tiles found "
              "(expected 'Faction BIO', 'Faction UCM', etc.)")
        sys.exit(1)

    # -- 5. Collect existing GUIDs --
    existing_guids = collect_all_guids(save_data)

    # -- 6. Process each ship row --
    groups: dict[str, list] = {}
    vectored_count  = 0
    monitored_count = 0

    for i, row in enumerate(rows):
        pdf_id    = str(row.get('PDF ID', 'UNKNOWN')).strip()
        ship_name = str(row['Name'])
        ship_type = str(row.get('Type', ''))
        card_url  = str(row.get('Card Image', ''))
        base_size = to_num(row['Size'])

        is_vectored, is_monitored = parse_special_flags(row)
        if is_vectored:  vectored_count  += 1
        if is_monitored: monitored_count += 1

        # a) Build per-ship variables
        ship_vars = build_ship_vars(row)

        # b) Build stripped ship JSON + state
        ship_obj_json, ship_state = build_stripped_ship_obj(
            template_obj, row, existing_guids
        )

        # c) Build the card entry (compact format)
        card_entry = {
            "name":        ship_name,
            "description": ship_type,
            "imageURL":    card_url,
            "tonnage":     detect_class_tonnage(ship_name, ship_type),
            "shipTonnage": detect_ship_tonnage(base_size),
            "baseScale":   1,
            "shipVars":    ship_vars,
            "shipState":   ship_state,
            "shipObjJson": ship_obj_json,
        }

        groups.setdefault(pdf_id, []).append(card_entry)

        if (i + 1) % 50 == 0:
            print(f"  Processed {i + 1}/{len(rows)} ships...")

    total_ships = len(rows)
    print(f"  Created {total_ships} ship entries across {len(groups)} factions.")
    print(f"  Special flags detected:")
    print(f"    isVectored  = true : {vectored_count} ships")
    print(f"    isMonitored = true : {monitored_count} ships")
    print(f"    Neither flag       : {total_ships - vectored_count - monitored_count} ships")

    # -- 7. Sort each group alphabetically --
    for pdf_id in groups:
        groups[pdf_id].sort(key=lambda c: c['name'].lower())

    # -- 8. Embed card data into each faction container tile --
    for pdf_id, card_entries in groups.items():
        if pdf_id not in faction_tiles:
            print(f"  WARNING: No container tile found for PDF ID '{pdf_id}' "
                  f"({len(card_entries)} ships orphaned)")
            continue

        tile_idx  = faction_tiles[pdf_id]
        container = save_data['ObjectStates'][tile_idx]
        lua_script = container['LuaScript']

        # Remove any existing embedded data section
        for marker in ("\r\n--EMBEDDED_DATA_START", "\n--EMBEDDED_DATA_START"):
            pos = lua_script.rfind(marker)
            if pos != -1:
                lua_script = lua_script[:pos]
                break

        # Build new embedded data
        embedded = build_embedded_data(card_entries, script_template)
        lua_script = lua_script + embedded

        container['LuaScript'] = lua_script
        print(f"  Embedded {len(card_entries)} ships into '{container['Nickname']}' "
              f"({len(embedded):,} chars)")

    # -- 9. Remove the template object from the save --
    if template_idx is not None:
        save_data['ObjectStates'] = [
            obj for i, obj in enumerate(save_data['ObjectStates'])
            if i != template_idx
        ]

    # -- 10. Write output --
    print(f"\nWriting output to: {output_path}")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(save_data, f, ensure_ascii=False)

    print("Done!")
    print(f"  Total objects on table: {len(save_data['ObjectStates'])}")


if __name__ == '__main__':
    main()
