import fitz  # PyMuPDF
from PIL import Image
import io
import re
import os
import numpy as np
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from scipy import ndimage

def extract_ship_name(page):
    text = page.get_text("text")
    lines = [line.strip() for line in text.split('\n') if line.strip()]

    skip_words = {
        'Cost','Effect','Special','Name','Arc','Att','Lock','DMG','Type',
        'Load','Launch','Thrust','Scan','Sig','Hull','ES','KS','BS','G'
    }

    def is_header(line):
        for word in skip_words:
            if line == word or line.startswith(word + ' ') or line.startswith(word + '\t'):
                return True
        return False

    for i, line in enumerate(lines):
        if is_header(line):
            continue

        l = " ".join(line.split())
        l_lower = l.lower()

        # If this line contains "570 pts" etc
        if re.search(r'\b\d+\s*pts\b', l_lower):

            combined = l

            # CASE A: points line starts with digits, name is probably the line above
            # e.g. "570 pts (45 + 525 pts)" with "Magellan - Coloniser" above it
            if re.match(r'^\d+\s*pts\b', l_lower) and i > 0:
                prev = " ".join(lines[i - 1].split())
                if prev and (not is_header(prev)):
                    combined = prev + " " + combined

            # CASE B: bracket is on the next line, append it
            if '(' not in combined and i + 1 < len(lines):
                nxt = " ".join(lines[i + 1].split())
                if nxt.startswith('('):
                    combined = combined + " " + nxt

            # Extract everything before the points value
            m = re.match(r'^(.+?)\s+(\d+)\s*pts\b', combined, re.IGNORECASE)
            if m:
                full_name = m.group(1).strip()
                full_name = re.sub(r'[<>:"/\\|?*]', '', full_name)
                return full_name

    return "Unknown_Ship"

def get_unique_filepath(filepath):
    """
    Check if a file exists, and if so, return an overflow filename instead.
    
    Args:
        filepath: Desired file path
    
    Returns:
        str: Either the original filepath or an overflow version
    """
    if not os.path.exists(filepath):
        return filepath
    
    # File exists, create overflow version
    directory = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    name, ext = os.path.splitext(filename)
    
    # Find the next available overflow number
    counter = 1
    while True:
        overflow_name = f"Overflow_{counter}{ext}"
        overflow_path = os.path.join(directory, overflow_name)
        if not os.path.exists(overflow_path):
            print(f"    ⚠ Duplicate detected: Renaming to '{overflow_name}'")
            return overflow_path
        counter += 1

def extract_ship_data(page):
    """
    Extract all ship data from a page for Excel export.
    
    Returns:
        dict: Dictionary with ship stats, or None if no ship data found
    """
    text = page.get_text("text")
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    # Check if this is a valid ship page (has enough lines)
    if len(lines) < 10:
        return None
    
    try:
        # Extract basic info
        ship_name = lines[0].strip()
        
        # If the name is "Load", treat it as no ship
        if ship_name.lower() == "load":
            return None
        
        # Extract points (look for "XX pts" pattern in first few lines)
        points = None
        for line in lines[:5]:
            points_match = re.search(r'(\d+)\s*pts', line, re.IGNORECASE)
            if points_match:
                points = int(points_match.group(1))
                break
        
        # Extract ship type (usually around line 2)
        ship_type = ""
        for line in lines[1:5]:
            # Skip the points line and size line
            if 'pts' not in line.lower() and '/' not in line and 'mm' not in line.lower():
                ship_type = line.strip()
                break
        
        # Extract size (look for "L/XXmm" or "L / XXmm" pattern)
        size = None
        for line in lines[:10]:
            size_match = re.search(r'[A-Z]\s*/\s*(\d+)\s*mm', line, re.IGNORECASE)
            if size_match:
                size = int(size_match.group(1))
                break
        
        # Find where "Thrust" header starts
        thrust_index = None
        for i, line in enumerate(lines):
            if line.strip().lower() == 'thrust':
                thrust_index = i
                break
        
        if thrust_index is None:
            return None
        
        # Map the headers to find where each stat value will be
        # Starting from thrust_index, we have: Thrust, Scan, Sig, Hull, ES, KS, BS, G, Special
        headers = ['Thrust', 'Scan', 'Sig', 'Hull', 'ES', 'KS', 'BS', 'G', 'Special']
        header_count = 0
        
        # Count how many consecutive header lines we have
        for i in range(thrust_index, min(thrust_index + 15, len(lines))):
            if lines[i].strip() in headers:
                header_count += 1
            else:
                break
        
        # The values start right after the headers
        values_start = thrust_index + header_count
        
        # Extract the values in order
        thrust = ""
        scan = ""
        sig = ""
        hull = ""
        es = ""
        ks = ""
        bs = ""
        g = ""
        special = ""
        
        # Helper function to extract only digits from a string
        def extract_digits(text):
            # Extract only digits from the text
            digits = re.sub(r'[^\d]', '', text)
            return digits if digits else ""
        
        # Read values in the same order as headers
        value_index = values_start
        
        if value_index < len(lines):
            # Thrust (e.g., "10"") - extract only digits
            thrust = extract_digits(lines[value_index].strip())
            value_index += 1
        
        if value_index < len(lines):
            # Scan (e.g., "6"") - extract only digits
            scan = extract_digits(lines[value_index].strip())
            value_index += 1
        
        if value_index < len(lines):
            # Sig (e.g., "3"") - extract only digits
            sig = extract_digits(lines[value_index].strip())
            value_index += 1
        
        if value_index < len(lines):
            # Hull (e.g., "4")
            hull = extract_digits(lines[value_index].strip())
            value_index += 1
        
        if value_index < len(lines):
            # ES (e.g., "4+") - extract only digits
            es = extract_digits(lines[value_index].strip())
            value_index += 1
        
        if value_index < len(lines):
            # KS (e.g., "3+") - extract only digits
            ks = extract_digits(lines[value_index].strip())
            value_index += 1
        
        if value_index < len(lines):
            # BS (e.g., "-" or a number)
            bs = lines[value_index].strip()
            value_index += 1
        
        if value_index < len(lines):
            # G (e.g., "2-4" or "1")
            g = lines[value_index].strip()
            value_index += 1
        
        if value_index < len(lines):
            # Special (might be multiple words, get everything remaining on this line and next few)
            special_parts = []
            for i in range(value_index, min(value_index + 5, len(lines))):
                # Stop if we hit the next section or empty-ish content
                if lines[i].strip() and not lines[i].strip().startswith('Name'):
                    special_parts.append(lines[i].strip())
                else:
                    break
            special = ', '.join(special_parts) if special_parts else ""
            
            # If "Famous" appears in special, only keep everything before it
            if 'Famous' in special:
                special = special.split('Famous')[0].strip()
                # Remove trailing comma if present
                special = special.rstrip(',').strip()
        
        return {
            'name': ship_name,
            'points': points if points else "",
            'type': ship_type,
            'size': size if size else "",
            'thrust': thrust,
            'scan': scan,
            'sig': sig,
            'hull': hull,
            'es': es,
            'ks': ks,
            'bs': bs,
            'g': g,
            'special': special.strip(),
            'pdf_id': '',  # Will be filled in by pdf_to_png
            'model_image': '',  # Will be filled in by pdf_to_png
            'card_image': ''  # Will be filled in by pdf_to_png
        }
        
    except Exception as e:
        print(f"    Error extracting ship data: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def get_crop_rect(page, search_texts=None, crop_whitespace=True):
    """
    Find text on the page and/or detect content boundaries to crop whitespace.
    
    Args:
        page: PyMuPDF page object
        search_texts: List of texts to search for cropping (searches all, uses earliest)
        crop_whitespace: Whether to crop empty whitespace at bottom
    
    Returns:
        fitz.Rect: Clipping rectangle, or None if no cropping needed
    """
    if search_texts is None:
        search_texts = [
            "Famous Ships of the class",
            "Known ships of the class", 
            "Recorded ships of the class",
            "Encountered ships of the class"
        ]
    
    page_rect = page.rect
    crop_y = page_rect.height  # Default to full page height
    
    # Search for all text patterns and find the earliest (lowest y-coordinate)
    earliest_y = page_rect.height
    found_text = None
    
    for search_text in search_texts:
        text_instances = page.search_for(search_text)
        if text_instances:
            # Get the y-coordinate where this text starts
            text_y = text_instances[0].y0
            if text_y < earliest_y:
                earliest_y = text_y
                found_text = search_text
    
    if found_text:
        crop_y = earliest_y
        # Optional: print which text was found for debugging
        # print(f"    Found '{found_text}' - cropping there")
    
    # If crop_whitespace is enabled, find the actual content boundary
    if crop_whitespace:
        # Get all text blocks with their positions
        blocks = page.get_text("dict")["blocks"]
        
        if blocks:
            # Find the lowest y-coordinate of any content
            max_y = 0
            for block in blocks:
                if "bbox" in block:
                    block_bottom = block["bbox"][3]  # y1 coordinate
                    max_y = max(max_y, block_bottom)
            
            # Add a small margin (e.g., 20 points) below the last content
            content_bottom = max_y + 20
            
            # Use the smaller of the two crop points
            crop_y = min(crop_y, content_bottom)
    
    # Only return a crop rectangle if we're actually cropping
    if crop_y < page_rect.height:
        clip_rect = fitz.Rect(0, 0, page_rect.width, crop_y)
        return clip_rect
    
    return None

def remove_background(image, threshold=245, black_threshold=15):
    """
    Remove white and black backgrounds from an image using edge-based flood fill.
    Only removes background pixels connected to edges, preserving white/black in the artwork.
    
    Args:
        image: PIL Image object
        threshold: Pixel values above this are considered white background (default: 245)
        black_threshold: Pixel values below this are considered black background (default: 15)
    
    Returns:
        PIL Image with transparent background
    """
    # Convert to RGBA if not already
    if image.mode != 'RGBA':
        image = image.convert('RGBA')
    
    # Convert to numpy array for easier manipulation
    data = np.array(image)
    
    # Get RGB channels
    r, g, b, a = data[:,:,0], data[:,:,1], data[:,:,2], data[:,:,3]
    
    # Create initial mask for potential background pixels
    white_candidates = (r > threshold) & (g > threshold) & (b > threshold)
    black_candidates = (r < black_threshold) & (g < black_threshold) & (b < black_threshold)
    
    # Create mask for pixels to make transparent (edge-connected only)
    height, width = data.shape[:2]
    background_mask = np.zeros((height, width), dtype=bool)
    
    # Flood fill from all edges to find connected background pixels
    from scipy import ndimage
    
    # Process white background
    if np.any(white_candidates):
        # Create a binary image of white candidates
        white_binary = white_candidates.astype(np.uint8)
        
        # Mark edge pixels as seeds
        seeds = np.zeros_like(white_binary)
        seeds[0, :] = white_binary[0, :]      # Top edge
        seeds[-1, :] = white_binary[-1, :]    # Bottom edge
        seeds[:, 0] = white_binary[:, 0]      # Left edge
        seeds[:, -1] = white_binary[:, -1]    # Right edge
        
        # Flood fill to find all white pixels connected to edges
        filled = ndimage.binary_dilation(seeds, mask=white_binary, iterations=-1)
        background_mask |= filled
    
    # Process black background
    if np.any(black_candidates):
        # Create a binary image of black candidates
        black_binary = black_candidates.astype(np.uint8)
        
        # Mark edge pixels as seeds
        seeds = np.zeros_like(black_binary)
        seeds[0, :] = black_binary[0, :]      # Top edge
        seeds[-1, :] = black_binary[-1, :]    # Bottom edge
        seeds[:, 0] = black_binary[:, 0]      # Left edge
        seeds[:, -1] = black_binary[:, -1]    # Right edge
        
        # Flood fill to find all black pixels connected to edges
        filled = ndimage.binary_dilation(seeds, mask=black_binary, iterations=-1)
        background_mask |= filled
    
    # Set alpha channel to 0 (transparent) for background pixels
    data[background_mask, 3] = 0
    
    # Convert back to PIL Image
    return Image.fromarray(data)

def extract_ship_image(page, ship_name, output_folder, dpi=300):
    """
    Extract the ship image from the page, ignoring large background images.
    Remove its background and save it with transparency.
    Optionally flip image horizontally for specific ships.
    
    Args:
        page: PyMuPDF page object
        ship_name: Name for the output file
        output_folder: Folder to save the image
        dpi: Resolution for extraction
    
    Returns:
        bool: True if image was extracted successfully
    """
    # List of ships that need horizontal flip
    ships_to_flip = [
        'Jakarta', 'London', 'Lysander', 'Taipei', 'Tayne', 'Warsaw',
        'L-Type', 'Nirvana', 'Precedent',
        'Aquamarine', 'Azurite', 'Basalt', 'Caesium', 'Granite', 'Jet',
        'Onyx', 'Turquoise', 'Uranium','Orion','Medea','Europa','Pandora',
        'Ajax','Perseus','Achilles','Hector','Orpheus','Ganymede','Romulus','Remus',
        'Cthulhu','Gargoyle','Gremlin','Harpy','Hydra','Nosferatu','Raiju','Scylla',
        'Sphinx','Strix','Succubus','Yokai','Amethyst','Plutonium','Halsey','Osaka',
        'Rio','Charybdis','Chimera','Djinn'
    ]
    
    # Check if this ship needs to be flipped
    needs_flip = any(flip_name in ship_name for flip_name in ships_to_flip)
    
    # Get all images on the page
    image_list = page.get_images(full=True)
    
    if not image_list:
        return False
    
    # Find the largest non-background image
    largest_image = None
    largest_area = 0
    largest_xref = None
    
    for img in image_list:
        xref = img[0]
        
        # Get the actual image dimensions
        try:
            base_image = page.parent.extract_image(xref)
            width = base_image["width"]
            height = base_image["height"]
            
            # Skip large background images (both dimensions > 2000 pixels)
            if width > 2000 and height > 2000:
                continue
            
            # Calculate area
            area = width * height
            
            # Track the largest non-background image
            if area > largest_area:
                largest_area = area
                largest_xref = xref
                
        except Exception as e:
            continue
    
    if not largest_xref:
        return False
    
    # Extract the ship image
    try:
        base_image = page.parent.extract_image(largest_xref)
        image_bytes = base_image["image"]
        
        # Load image with PIL
        image = Image.open(io.BytesIO(image_bytes))
        
        # Apply flip BEFORE background removal if needed
        if needs_flip:
            image = image.transpose(Image.Transpose.FLIP_LEFT_RIGHT)
            # Force a copy to ensure the flip is preserved
            image = image.copy()
            print(f"    → Flipped horizontally")
        
        # Remove background and make transparent
        image_transparent = remove_background(image)
        
        # Save the image with duplicate protection
        output_filename = f"{ship_name}_ModelImage.png"
        output_path = os.path.join(output_folder, output_filename)
        output_path = get_unique_filepath(output_path)
        
        image_transparent.save(output_path, "PNG")
        
        return True
        
    except Exception as e:
        print(f"    Warning: Could not extract ship image - {str(e)}")
        return False

def pdf_to_png(pdf_path, output_folder="output", dpi=300, crop_text="Famous Ships of the class", 
               crop_whitespace=True, extract_model=True, ship_data_list=None):
    """
    Convert each page of a PDF to PNG with ship name in filename.
    Crops content at specified text and/or removes bottom whitespace.
    Optionally extracts ship model images with transparent backgrounds.
    
    Args:
        pdf_path: Path to the input PDF file
        output_folder: Folder to save PNG files (default: 'output')
        dpi: Resolution for the output images (default: 300)
        crop_text: Text to search for cropping (default: 'Famous Ships of the class')
        crop_whitespace: Whether to crop empty whitespace at bottom (default: True)
        extract_model: Whether to extract ship model images (default: True)
        ship_data_list: List to append ship data to for Excel export
    """
    # Create output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Open the PDF
    pdf_document = fitz.open(pdf_path)
    
    # Extract first 3 letters of PDF filename (without path and extension)
    pdf_basename = os.path.splitext(os.path.basename(pdf_path))[0]
    pdf_identifier = pdf_basename[:3].upper()  # First 3 letters, uppercase
    
    print(f"Processing PDF: {pdf_path}")
    print(f"Total pages: {len(pdf_document)}")
    print(f"PDF Identifier: {pdf_identifier}")
    
    # Process each page
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]
        
        # Extract ship name from the page
        ship_name = extract_ship_name(page)
        
        # Create the filenames
        model_filename = f"{ship_name}_ModelImage.png"
        card_filename = f"{ship_name}_CardFrontImage.png"
        
        # Create full GitHub URLs
        base_url = "https://raw.githubusercontent.com/TemporalDistoriton/DropfleetTTS/main/RemasterShips"
        model_url = f"{base_url}/{pdf_identifier}/{model_filename}"
        card_url = f"{base_url}/{pdf_identifier}/{card_filename}"
        
        # Extract ship data for Excel
        ship_data = extract_ship_data(page)
        if ship_data and ship_data_list is not None:
            # Add the PDF identifier and full URLs to the ship data
            ship_data['pdf_id'] = pdf_identifier
            ship_data['model_image'] = model_url
            ship_data['card_image'] = card_url
            ship_data_list.append(ship_data)
        elif ship_data_list is not None:
            # No ship data found
            ship_data_list.append({
                'name': 'no ship',
                'pdf_id': pdf_identifier,
                'model_image': '',
                'card_image': ''
            })
        
        # Create the output path for card front with duplicate protection
        output_path = os.path.join(output_folder, card_filename)
        output_path = get_unique_filepath(output_path)
        
        # Calculate zoom factor based on DPI (72 is the default PDF DPI)
        zoom = dpi / 72
        mat = fitz.Matrix(zoom, zoom)
        
        # Check if we need to crop the page
        clip_rect = get_crop_rect(page, crop_whitespace=crop_whitespace)
        
        if clip_rect:
            # Render only the cropped portion
            pix = page.get_pixmap(matrix=mat, clip=clip_rect)
            print(f"Page {page_num + 1}: Cropped - Saved as '{os.path.basename(output_path)}'")
        else:
            # Render the full page
            pix = page.get_pixmap(matrix=mat)
            print(f"Page {page_num + 1}: Full page - Saved as '{os.path.basename(output_path)}'")
        
        # Save card front as PNG with error handling
        try:
            pix.save(output_path)
        except Exception as e:
            print(f"    ⚠ Cannot save card: {e}")
            overflow_path = os.path.join(output_folder, f"Overflow_Card_{page_num}.png")
            pix.save(overflow_path)
            print(f"    → Saved as '{os.path.basename(overflow_path)}' instead")
        
        # Extract ship model image if requested
        if extract_model:
            success = extract_ship_image(page, ship_name, output_folder, dpi)
            if success:
                print(f"    ✓ Extracted model image")
            else:
                print(f"    ✗ No model image found")
    
    pdf_document.close()
    print(f"\nConversion complete! Files saved in '{output_folder}' folder.")

def create_excel_file(ship_data_list, output_file="ship_data.xlsx"):
    """
    Create an Excel file with all ship data.
    
    Args:
        ship_data_list: List of ship data dictionaries
        output_file: Name of the output Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ship Data"
    
    # Define headers - PDF ID, Model Image, Card Image at the end
    headers = ['Name', 'Points', 'Type', 'Size', 'Thrust', 'Scan', 'Sig', 
               'Hull', 'ES', 'KS', 'BS', 'G', 'Special', 'PDF ID', 'Model Image', 'Card Image']
    
    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, ship in enumerate(ship_data_list, start=2):
        # Always write these columns
        ws.cell(row=row_idx, column=1, value=ship.get('name', ''))
        ws.cell(row=row_idx, column=14, value=ship.get('pdf_id', ''))
        ws.cell(row=row_idx, column=15, value=ship.get('model_image', ''))
        ws.cell(row=row_idx, column=16, value=ship.get('card_image', ''))
        
        # Write other columns only if not "no ship"
        if ship.get('name') != 'no ship':
            ws.cell(row=row_idx, column=2, value=ship.get('points', ''))
            ws.cell(row=row_idx, column=3, value=ship.get('type', ''))
            ws.cell(row=row_idx, column=4, value=ship.get('size', ''))
            ws.cell(row=row_idx, column=5, value=ship.get('thrust', ''))
            ws.cell(row=row_idx, column=6, value=ship.get('scan', ''))
            ws.cell(row=row_idx, column=7, value=ship.get('sig', ''))
            ws.cell(row=row_idx, column=8, value=ship.get('hull', ''))
            ws.cell(row=row_idx, column=9, value=ship.get('es', ''))
            ws.cell(row=row_idx, column=10, value=ship.get('ks', ''))
            ws.cell(row=row_idx, column=11, value=ship.get('bs', ''))
            ws.cell(row=row_idx, column=12, value=ship.get('g', ''))
            ws.cell(row=row_idx, column=13, value=ship.get('special', ''))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved: {output_file}")

# Example usage
if __name__ == "__main__":
    import sys
    
    # DPI setting (can be adjusted here if needed)
    dpi = 300  # Higher DPI = better quality but larger files
    
    # Find all PDF files in the current directory
    pdf_files = glob.glob("*.pdf")
    
    if not pdf_files:
        print("No PDF files found in the current directory!")
        print("Please make sure there are PDF files in the same folder as this script.")
        sys.exit(1)
    
    print(f"Found {len(pdf_files)} PDF file(s) to process:")
    for pdf in pdf_files:
        print(f"  - {pdf}")
    print()
    
    # List to collect all ship data for Excel
    all_ship_data = []
    
    # Process each PDF file
    for pdf_file in pdf_files:
        # Create output folder name from PDF filename (without extension)
        pdf_basename = os.path.splitext(os.path.basename(pdf_file))[0]
        output_folder = pdf_basename[:3].upper()  # First 3 letters, uppercase
        
        try:
            pdf_to_png(pdf_file, output_folder, dpi, ship_data_list=all_ship_data)
            print()
        except Exception as e:
            print(f"Error processing {pdf_file}: {str(e)}")
            print()
    
    # Create Excel file with all ship data
    if all_ship_data:
        create_excel_file(all_ship_data, "ship_data.xlsx")
    
    print("=" * 60)
    print("All PDFs processed!")